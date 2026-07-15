"""Nomes da turma, temas e mapeamento por ficheiro importado."""

from __future__ import annotations

import io
import json
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

from app import db
from app.config import NOMES_ALUNOS_PATH

if TYPE_CHECKING:
    from app.models import AlunoRelatorio

# Conectores ignorados ao gerar a chave do ficheiro a partir do nome.
_CONECTORES_NOME = {"de", "da", "do", "das", "dos", "e", "di", "du"}

_KV_NOMES = "nomes_turma"


@dataclass
class AlunoTurma:
    nome: str
    chave_ficheiro: str = ""
    tema: str = ""


def _sem_acentos(texto: str) -> str:
    norm = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in norm if not unicodedata.combining(c))


def gerar_chave_ficheiro(nome: str) -> str:
    """Gera a chave do ficheiro a partir do nome: PrimeiroÚltimo, sem acentos/espaços.

    Ex.: "Adriano Ricardo David Salucombo" -> "AdrianoSalucombo".
    """
    tokens = [t for t in nome.strip().split() if t]
    if not tokens:
        return ""
    principais = [t for t in tokens if t.lower() not in _CONECTORES_NOME] or tokens
    escolhidos = [principais[0]] if len(principais) == 1 else [principais[0], principais[-1]]
    partes = []
    for token in escolhidos:
        limpo = "".join(c for c in _sem_acentos(token) if c.isalnum())
        if limpo:
            partes.append(limpo[:1].upper() + limpo[1:].lower())
    return "".join(partes)


def _parse_nomes(dados: list) -> list[AlunoTurma]:
    return [
        AlunoTurma(
            nome=str(item.get("nome", "")).strip(),
            chave_ficheiro=str(item.get("chave_ficheiro", "")).strip(),
            tema=str(item.get("tema", "")).strip(),
        )
        for item in dados
        if str(item.get("nome", "")).strip()
    ]


def carregar_nomes_turma() -> list[AlunoTurma]:
    valor = db.kv_get(_KV_NOMES)
    if valor is not None:
        return _parse_nomes(json.loads(valor))
    if NOMES_ALUNOS_PATH.exists():
        alunos = _parse_nomes(json.loads(NOMES_ALUNOS_PATH.read_text(encoding="utf-8")))
        guardar_nomes_turma(alunos)
        return alunos
    return []


def guardar_nomes_turma(alunos: list[AlunoTurma]) -> None:
    payload = [
        {
            "nome": a.nome.strip(),
            "chave_ficheiro": a.chave_ficheiro.strip(),
            "tema": a.tema.strip(),
        }
        for a in alunos
        if a.nome.strip()
    ]
    db.kv_set(_KV_NOMES, json.dumps(payload, ensure_ascii=False))


def tema_para_nome(nome: str) -> str:
    for aluno in carregar_nomes_turma():
        if aluno.nome == nome:
            return aluno.tema
    return ""


def _norm_cabecalho(valor: object) -> str:
    return _sem_acentos(str(valor or "")).strip().lower()


_COLUNAS_NOME = {"nome", "nome completo", "aluno", "nome do aluno"}
_COLUNAS_TEMA = {"tema", "tema pap", "tema da pap", "titulo", "titulo pap", "projeto"}
_COLUNAS_CHAVE = {"chave", "chave no ficheiro", "chave ficheiro", "chave do ficheiro"}


def importar_nomes_de_excel(conteudo: bytes) -> list[AlunoTurma]:
    """Lê um Excel com colunas Nome e Tema e devolve a lista da turma.

    A chave do ficheiro é gerada automaticamente a partir do nome, exceto se
    o Excel já tiver uma coluna de chave preenchida.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    ws = wb.active
    linhas = [linha for linha in ws.iter_rows(values_only=True) if linha]
    if not linhas:
        return []

    cabecalho = [_norm_cabecalho(c) for c in linhas[0]]
    idx_nome = idx_tema = idx_chave = None
    for i, titulo in enumerate(cabecalho):
        if idx_nome is None and titulo in _COLUNAS_NOME:
            idx_nome = i
        elif idx_tema is None and titulo in _COLUNAS_TEMA:
            idx_tema = i
        elif idx_chave is None and titulo in _COLUNAS_CHAVE:
            idx_chave = i

    if idx_nome is None:
        # Sem cabeçalho reconhecido: assume coluna 0 = nome, coluna 1 = tema.
        idx_nome, idx_tema = 0, 1
        dados = linhas
    else:
        dados = linhas[1:]

    def _cel(linha, idx):
        if idx is None or idx >= len(linha):
            return ""
        return str(linha[idx] or "").strip()

    alunos: list[AlunoTurma] = []
    for linha in dados:
        nome = _cel(linha, idx_nome)
        if not nome:
            continue
        chave = _cel(linha, idx_chave) or gerar_chave_ficheiro(nome)
        alunos.append(
            AlunoTurma(nome=nome, chave_ficheiro=chave, tema=_cel(linha, idx_tema))
        )
    return alunos


def colunas_turma_ordenadas(
    alunos: list[AlunoRelatorio],
) -> list[tuple[str, AlunoRelatorio | None]]:
    por_nome = {a.nome: a for a in alunos}
    return [
        (entry.nome, por_nome.get(entry.nome))
        for entry in carregar_nomes_turma()
    ]


def _corresponde_chave(stem: str, chave: str) -> bool:
    chave = chave.strip()
    if not chave:
        return False
    return chave.lower() in stem.lower()


def nome_por_ficheiro(
    nome_ficheiro: str,
    alunos: list[AlunoTurma] | None = None,
) -> str | None:
    if alunos is None:
        alunos = carregar_nomes_turma()
    stem = Path(nome_ficheiro).stem
    for entry in alunos:
        if _corresponde_chave(stem, entry.chave_ficheiro):
            return entry.nome
    return None


def aplicar_nomes_a_alunos(storage) -> tuple[int, list[str]]:
    """Actualiza nomes/temas dos alunos importados com base na configuração."""
    alunos_cfg = carregar_nomes_turma()
    actualizados = 0
    avisos: list[str] = []
    for aluno in storage.listar_alunos():
        novo_nome = nome_por_ficheiro(aluno.ficheiro, alunos_cfg)
        if not novo_nome:
            avisos.append(f"Sem mapeamento: {aluno.ficheiro}")
            continue
        novo_tema = tema_para_nome(novo_nome)
        if novo_nome != aluno.nome or (novo_tema and novo_tema != aluno.tema_pap):
            storage.atualizar_aluno(
                aluno.id,
                nome=novo_nome,
                tema_pap=novo_tema or aluno.tema_pap,
            )
            actualizados += 1
    return actualizados, avisos
