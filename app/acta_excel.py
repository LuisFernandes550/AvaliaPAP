"""Exportação completa para o ficheiro oficial Acta_Pap2526.xlsx."""

from __future__ import annotations

import re
import shutil
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.config import ACTA_MODELO_PATH, ACTA_PATH
from app.models import CriterioAvaliacao
from app.temas_pap import tema_para_nome

if TYPE_CHECKING:
    from app.models import AlunoRelatorio

FOLHA_AVALIACAO = "Avaliação"
FOLHA_AUXILIAR = "Auxiliar"
FOLHA_ACTA = "Acta"
FOLHAS_FIXAS = {FOLHA_AUXILIAR, FOLHA_ACTA}

LINHA_NOMES = 3
COLUNA_NOMES_AUXILIAR = 2
COLUNA_TEMAS_AUXILIAR = 4
PRIMEIRA_LINHA_ALUNO_AUXILIAR = 24
PRIMEIRA_COLUNA_ALUNO = 10

LINHAS_CRITERIOS = {
    15: CriterioAvaliacao.DEFINICAO_OBJETIVOS,
    16: CriterioAvaliacao.QUALIDADE_CIENTIFICA,
    17: CriterioAvaliacao.PERTINENCIA_CRIATIVIDADE,
    18: CriterioAvaliacao.PLANEAMENTO_RECURSOS,
    19: CriterioAvaliacao.FASES_EXECUCAO,
    20: CriterioAvaliacao.AUTONOMIA,
    21: CriterioAvaliacao.RESPONSABILIDADE_PROJETO,
    24: CriterioAvaliacao.OBJETIVIDADE,
    25: CriterioAvaliacao.PERTINENCIA,
    26: CriterioAvaliacao.DIFICULDADES,
    27: CriterioAvaliacao.ANALISE_CRITICA,
    28: CriterioAvaliacao.RESPONSABILIDADE_RELATORIO,
    31: CriterioAvaliacao.EXPRESSAO_ORAL,
    32: CriterioAvaliacao.CAPACIDADE_SINTESE,
    33: CriterioAvaliacao.RECURSOS_APRESENTACAO,
    34: CriterioAvaliacao.ARGUMENTACAO_DEFESA,
}
LINHA_AUTOAVALIACAO = 39

ANO_LETIVO = "Ano letivo 2025/2026"
LINHA_ANO = 4
LINHA_ALUNO = 8
LINHA_TEMA = 9
LINHA_AREA = 10
LINHA_CURSO = 11
LINHA_SAIDA = 12
COL_VALOR = 4

DEFAULT_AREA = "Ciências Informáticas (481)"
DEFAULT_CURSO = "Técnico de Gestão e Programação de Sistemas Informáticos"
DEFAULT_SAIDA = "Técnico de Gestão e Programação de Sistemas Informáticos"
CELULAS_VALIDACAO = frozenset({"D10", "D11", "D12"})


def garantir_acta() -> bool:
    """Copia o modelo da Acta para data/export/ se ainda não existir."""
    if ACTA_PATH.exists():
        return True
    if not ACTA_MODELO_PATH.exists():
        return False
    ACTA_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(ACTA_MODELO_PATH, ACTA_PATH)
    return True


def _repor_acta_modelo() -> bool:
    """Substitui a Acta de trabalho por uma cópia limpa do modelo."""
    if not ACTA_MODELO_PATH.exists():
        return garantir_acta()
    ACTA_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(ACTA_MODELO_PATH, ACTA_PATH)
    return True


def _normalizar_nome(nome: str) -> str:
    if not nome:
        return ""
    texto = unicodedata.normalize("NFKD", str(nome))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto.strip().casefold())


def _folha_avaliacao(wb):
    if FOLHA_AVALIACAO in wb.sheetnames:
        return wb[FOLHA_AVALIACAO]
    for nome in wb.sheetnames:
        if _normalizar_nome(nome).startswith("avaliacao"):
            return wb[nome]
    raise ValueError(f"Folha «{FOLHA_AVALIACAO}» não encontrada no ficheiro Acta.")


def _linhas_sync() -> dict[int, CriterioAvaliacao]:
    linhas = dict(LINHAS_CRITERIOS)
    linhas[LINHA_AUTOAVALIACAO] = CriterioAvaliacao.AUTOAVALIACAO
    return linhas


def _tema_aluno(aluno: AlunoRelatorio) -> str:
    if aluno.tema_pap:
        return aluno.tema_pap
    if aluno.titulo_pap:
        return aluno.titulo_pap
    return tema_para_nome(aluno.nome)


def _nome_folha_aluno(nome: str, usados: set[str]) -> str:
    partes = [p for p in nome.split() if p]
    candidatos: list[str] = []
    if len(partes) >= 2:
        candidatos.append(f"{partes[0]} {partes[-1]}")
    if len(partes) >= 3:
        candidatos.append(f"{partes[0]} {partes[1]}")
    if partes:
        candidatos.append(partes[0])
    for candidato in candidatos:
        titulo = candidato[:31]
        if titulo not in usados:
            usados.add(titulo)
            return titulo
    base = (candidatos[0] if candidatos else "Aluno")[:28]
    n = 2
    while True:
        titulo = f"{base} {n}"[:31]
        if titulo not in usados:
            usados.add(titulo)
            return titulo
        n += 1


def _folhas_aluno(wb) -> list[str]:
    nomes: list[str] = []
    depois_avaliacao = False
    for nome in wb.sheetnames:
        if not depois_avaliacao:
            if nome == FOLHA_AVALIACAO or _normalizar_nome(nome).startswith("avaliacao"):
                depois_avaliacao = True
            continue
        if nome in FOLHAS_FIXAS:
            break
        nomes.append(nome)
    return nomes


def _limpar_folha_auxiliar(ws, n_alunos: int) -> None:
    for linha in range(PRIMEIRA_LINHA_ALUNO_AUXILIAR, ws.max_row + 1):
        ws.cell(linha, COLUNA_NOMES_AUXILIAR, None)
        ws.cell(linha, COLUNA_TEMAS_AUXILIAR, None)
    ws.cell(22, COLUNA_NOMES_AUXILIAR, "Alunos")
    ws.cell(22, COLUNA_TEMAS_AUXILIAR, "Temas")


def _preencher_folha_auxiliar(ws, alunos: list[AlunoRelatorio]) -> None:
    _limpar_folha_auxiliar(ws, len(alunos))
    for indice, aluno in enumerate(alunos):
        linha = PRIMEIRA_LINHA_ALUNO_AUXILIAR + indice
        ws.cell(linha, COLUNA_NOMES_AUXILIAR, aluno.nome)
        ws.cell(linha, COLUNA_TEMAS_AUXILIAR, _tema_aluno(aluno))


def _nome_coluna_avaliacao(wb, ws, ws_valores, col: int) -> str | None:
    """Resolve o nome do aluno associado a uma coluna da folha Avaliação (linha 3)."""
    bruto = ws.cell(LINHA_NOMES, col).value
    if isinstance(bruto, str) and bruto.startswith("="):
        m = re.search(r"\$B(\d+)", bruto, re.IGNORECASE)
        if m and FOLHA_AUXILIAR in wb.sheetnames:
            nome = wb[FOLHA_AUXILIAR].cell(int(m.group(1)), COLUNA_NOMES_AUXILIAR).value
            if nome:
                return _normalizar_nome(str(nome))
    cached = ws_valores.cell(LINHA_NOMES, col).value
    if cached:
        return _normalizar_nome(str(cached))
    if bruto and not str(bruto).startswith("="):
        return _normalizar_nome(str(bruto))
    return None


def _mapear_colunas_avaliacao(wb, ws, ws_valores) -> dict[str, int]:
    """Mapa nome normalizado → índice de coluna na folha Avaliação."""
    mapa: dict[str, int] = {}
    ultima = max(ws.max_column, ws_valores.max_column, PRIMEIRA_COLUNA_ALUNO)
    for col in range(PRIMEIRA_COLUNA_ALUNO, ultima + 1):
        nome = _nome_coluna_avaliacao(wb, ws, ws_valores, col)
        if nome and nome not in mapa:
            mapa[nome] = col
    return mapa


def _limpar_notas_folha_avaliacao(ws, col_inicio: int, col_fim: int) -> None:
    """Remove notas existentes nas colunas de alunos antes de exportar."""
    linhas = _linhas_sync()
    for col in range(col_inicio, col_fim + 1):
        for linha in linhas:
            ws.cell(linha, col, None)


def _preencher_folha_avaliacao(
    wb,
    ws,
    ws_valores,
    alunos: list[AlunoRelatorio],
    avaliacoes: dict[int, dict],
) -> tuple[int, dict[int, int], list[str]]:
    """
    Localiza as colunas pelos nomes na linha 3 e actualiza apenas as notas
    nas células de critérios dessas colunas.
    Devolve (notas exportadas, mapa aluno_id→coluna, avisos).
    """
    linhas = _linhas_sync()
    mapa_nomes = _mapear_colunas_avaliacao(wb, ws, ws_valores)
    colunas_aluno: dict[int, int] = {}
    avisos: list[str] = []
    exportados = 0

    if not mapa_nomes:
        avisos.append("Nenhuma coluna com nome de aluno encontrada na folha Avaliação.")
        return 0, {}, avisos

    for aluno in alunos:
        chave = _normalizar_nome(aluno.nome)
        col = mapa_nomes.get(chave)
        if col is None:
            avisos.append(f"Coluna não encontrada na Avaliação para {aluno.nome}.")
            continue

        colunas_aluno[aluno.id] = col
        av = avaliacoes.get(aluno.id, {})
        for linha, criterio in linhas.items():
            if criterio in av:
                ws.cell(linha, col, int(av[criterio].nota))
                exportados += 1
            else:
                ws.cell(linha, col, None)

    return exportados, colunas_aluno, avisos


def _referencia_avaliacao(nome_folha: str, col_letra: str, linha: int, factor: int) -> str:
    folha = nome_folha.replace("'", "''")
    return f"='{folha}'!{col_letra}{linha}*{factor}"


def _refs_validacao_auxiliar(nome_folha_aux: str) -> tuple[str, str, str]:
    folha = nome_folha_aux.replace("'", "''")
    return (
        f"='{folha}'!$B$6:$B$8",
        f"='{folha}'!$D$6:$D$8",
        f"='{folha}'!$B$16:$B$18",
    )


def _celulas_validacao(sqref: str) -> set[str]:
    from openpyxl.utils.cell import range_boundaries

    celulas: set[str] = set()
    for parte in str(sqref).split():
        if ":" in parte:
            min_col, min_row, max_col, max_row = range_boundaries(parte)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    celulas.add(f"{get_column_letter(col)}{row}")
        else:
            celulas.add(parte)
    return celulas


def _remover_validacoes_celulas(ws, celulas: set[str]) -> None:
    restantes = [
        dv
        for dv in ws.data_validations.dataValidation
        if not (_celulas_validacao(str(dv.sqref)) & celulas)
    ]
    ws.data_validations.dataValidation = restantes


def _aplicar_validacao_formacao(ws, nome_folha_aux: str) -> None:
    _remover_validacoes_celulas(ws, set(CELULAS_VALIDACAO))
    for cell, formula in zip(
        CELULAS_VALIDACAO,
        _refs_validacao_auxiliar(nome_folha_aux),
    ):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Seleccione um valor da lista."
        dv.errorTitle = "Valor inválido"
        dv.add(cell)
        ws.add_data_validation(dv)


def _preencher_cabecalho_folha_aluno(
    ws,
    nome_folha_aux: str,
    aluno: AlunoRelatorio | None = None,
) -> None:
    ws.cell(LINHA_ANO, 1, ANO_LETIVO)
    ws.cell(LINHA_ALUNO, 1, "Aluno:")
    ws.cell(LINHA_TEMA, 3, "Tema:")
    ws.cell(LINHA_AREA, 1, "Área de Formação:")
    ws.cell(LINHA_CURSO, 1, "Curso Profissional:")
    ws.cell(LINHA_SAIDA, 1, "Saída Profissional:")

    if aluno is not None:
        ws.cell(LINHA_ALUNO, COL_VALOR, aluno.nome)
        ws.cell(LINHA_TEMA, COL_VALOR, _tema_aluno(aluno))

    for linha, default in (
        (LINHA_AREA, DEFAULT_AREA),
        (LINHA_CURSO, DEFAULT_CURSO),
        (LINHA_SAIDA, DEFAULT_SAIDA),
    ):
        if not ws.cell(linha, COL_VALOR).value:
            ws.cell(linha, COL_VALOR, default)

    _aplicar_validacao_formacao(ws, nome_folha_aux)


def _atualizar_folha_aluno(
    ws,
    aluno: AlunoRelatorio,
    col_letra: str,
    nome_folha_av: str,
    nome_folha_aux: str,
) -> None:
    _preencher_cabecalho_folha_aluno(ws, nome_folha_aux, aluno)
    ws.cell(16, 10, _referencia_avaliacao(nome_folha_av, col_letra, 22, 5))
    ws.cell(26, 10, _referencia_avaliacao(nome_folha_av, col_letra, 29, 2))
    ws.cell(34, 10, _referencia_avaliacao(nome_folha_av, col_letra, 35, 3))
    ws.cell(39, 10, "=ROUND(SUM(J16,J26,J34),0)")


def _mover_antes_auxiliar(wb, ws) -> None:
    if FOLHA_AUXILIAR not in wb.sheetnames:
        return
    dest = wb.sheetnames.index(FOLHA_AUXILIAR)
    actual = wb.sheetnames.index(ws.title)
    if actual != dest:
        wb.move_sheet(ws, offset=dest - actual)


def _sincronizar_folhas_aluno(
    wb,
    alunos: list[AlunoRelatorio],
    nome_folha_av: str,
    colunas_aluno: dict[int, int],
    nome_folha_aux: str,
) -> int:
    folhas_existentes = _folhas_aluno(wb)
    if folhas_existentes:
        template = wb[folhas_existentes[0]]
    elif FOLHA_ACTA in wb.sheetnames:
        template = wb[FOLHA_ACTA]
    else:
        raise ValueError("Não foi encontrado modelo para folhas individuais de aluno.")

    usados: set[str] = set()
    folhas_alvo: list[str] = []

    for indice, aluno in enumerate(alunos):
        titulo = _nome_folha_aluno(aluno.nome, usados)
        col = colunas_aluno.get(aluno.id)
        if col is None:
            continue
        col_letra = get_column_letter(col)

        if indice < len(folhas_existentes):
            nome_antigo = folhas_existentes[indice]
            ws = wb[nome_antigo]
            if ws.title != titulo:
                ws.title = titulo
        else:
            ws = wb.copy_worksheet(template)
            ws.title = titulo
            _mover_antes_auxiliar(wb, ws)

        _atualizar_folha_aluno(ws, aluno, col_letra, nome_folha_av, nome_folha_aux)
        folhas_alvo.append(titulo)

    for nome in folhas_existentes[len(alunos):]:
        if nome in wb.sheetnames:
            del wb[nome]

    for nome in _folhas_aluno(wb):
        if nome not in folhas_alvo:
            _preencher_cabecalho_folha_aluno(wb[nome], nome_folha_aux)

    if FOLHA_ACTA in wb.sheetnames:
        _preencher_cabecalho_folha_aluno(wb[FOLHA_ACTA], nome_folha_aux)

    return len(folhas_alvo)


@dataclass
class ResultadoSincronizacaoActa:
    importados: int
    exportados: int
    caminho: Path
    bytes_ficheiro: bytes
    avisos: list[str]


def sincronizar_acta(
    alunos: list[AlunoRelatorio],
    avaliacoes: dict[int, dict],
) -> ResultadoSincronizacaoActa:
    """
    Substitui o conteúdo da Acta pelos alunos e notas da aplicação:
    - Folha Auxiliar: lista de alunos e temas
    - Folha Avaliação: notas por critério
    - Folhas individuais: uma por aluno (reutilizando os separadores existentes)
    """
    avisos: list[str] = []
    if not _repor_acta_modelo():
        avisos.append(
            f"Ficheiro Acta não encontrado. Modelo em falta: {ACTA_MODELO_PATH}."
        )
        return ResultadoSincronizacaoActa(0, 0, ACTA_PATH, b"", avisos)

    alunos_ordem = [a for a in alunos if a.id is not None]
    if not alunos_ordem:
        avisos.append("Nenhum aluno na aplicação para exportar.")
        return ResultadoSincronizacaoActa(0, 0, ACTA_PATH, ACTA_PATH.read_bytes(), avisos)

    backup = ACTA_PATH.with_suffix(".xlsx.bak")
    shutil.copy2(ACTA_PATH, backup)

    wb = load_workbook(ACTA_PATH, keep_links=True)
    wb_valores = load_workbook(ACTA_PATH, data_only=True, keep_links=True)
    try:
        if FOLHA_AUXILIAR not in wb.sheetnames:
            raise ValueError(f"Folha «{FOLHA_AUXILIAR}» não encontrada.")

        ws_aux = wb[FOLHA_AUXILIAR]
        ws_av = _folha_avaliacao(wb)
        ws_av_valores = _folha_avaliacao(wb_valores)
        nome_folha_av = ws_av.title
        nome_folha_aux = ws_aux.title

        ultima_col = max(
            ws_av.max_column,
            ws_av_valores.max_column,
            PRIMEIRA_COLUNA_ALUNO,
        )
        _limpar_notas_folha_avaliacao(ws_av, PRIMEIRA_COLUNA_ALUNO, ultima_col)

        _preencher_folha_auxiliar(ws_aux, alunos_ordem)

        exportados, colunas_aluno, avisos_av = _preencher_folha_avaliacao(
            wb, ws_av, ws_av_valores, alunos_ordem, avaliacoes
        )
        avisos.extend(avisos_av)

        n_folhas = _sincronizar_folhas_aluno(
            wb, alunos_ordem, nome_folha_av, colunas_aluno, nome_folha_aux
        )

        ACTA_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(ACTA_PATH)
    except Exception:
        shutil.copy2(backup, ACTA_PATH)
        raise
    finally:
        wb.close()
        wb_valores.close()

    avisos.append(f"{len(alunos_ordem)} aluno(s) exportado(s).")
    avisos.append(f"{n_folhas} separador(es) de aluno actualizado(s).")

    return ResultadoSincronizacaoActa(
        importados=0,
        exportados=exportados,
        caminho=ACTA_PATH,
        bytes_ficheiro=ACTA_PATH.read_bytes(),
        avisos=avisos,
    )
