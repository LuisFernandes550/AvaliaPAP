from __future__ import annotations

import html
import io
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from st_aggrid import AgGrid, DataReturnMode, GridOptionsBuilder, JsCode

from app.ai_evaluator import avaliar_relatorio, ia_disponivel, resumir_capitulos
from app.acta_excel import sincronizar_acta
from app.app_settings import (
    TITULO_PADRAO,
    ConfiguracaoApp,
    carregar_configuracao_app,
    guardar_configuracao_app,
    titulo_app,
)
from app.auth import AuthStorage, Utilizador
from app.backup import descricao_armazenamento, exportar_backup, importar_backup
from app.config import ACTA_PATH, EM_STREAMLIT_CLOUD, NOTA_MAXIMA, RELATORIOS_DIR
from app.models import (
    AREA_LABELS,
    CAPITULOS_PAP,
    CRITERIO_LABELS,
    CRITERIO_MANUAL,
    CRITERIOS_MANUAIS,
    CRITERIOS_POR_SECAO,
    SECAO_LABELS,
    SECAO_PONDERACAO,
    AlunoRelatorio,
    AreaPAP,
    CriterioAvaliacao,
    ResultadoCriterio,
    SecaoAvaliacao,
    avaliacao_final,
    media_secao,
    media_secao_arredondada,
    nota_final_arredondada,
)
from app.pdf_converter import caminho_pdf_para_docx, docx_para_html, gerar_pdf
from app.report_parser import analisar_relatorio
from app.temas_pap import colunas_turma_ordenadas, tema_para_nome
from app.nomes_alunos import (
    AlunoTurma,
    aplicar_nomes_a_alunos,
    carregar_nomes_turma,
    guardar_nomes_turma,
    nome_por_ficheiro,
    nomes_turma_default,
)
from app.storage import (
    PapStorage,
    carregar_instrucoes,
    guardar_instrucoes,
    instrucoes_default,
)

st.set_page_config(page_title=titulo_app(), page_icon="🎓", layout="wide")

st.markdown(
    """
    <style>
    [data-testid="InputInstructions"] { display: none !important; }
    [data-testid="stToolbar"] { display: none !important; }
    .stAppDeployButton { display: none !important; }
    header[data-testid="stHeader"] {
        height: 0 !important;
        min-height: 0 !important;
        visibility: hidden !important;
        display: none !important;
    }
    [data-testid="stAppViewContainer"] > section > div {
        padding-top: 0 !important;
    }
    [data-testid="stMain"] {
        padding-top: 0 !important;
    }
    section.main > div.block-container {
        padding-top: 2rem !important;
        padding-bottom: 1rem !important;
        max-width: 100%;
    }
    .st-emotion-cache-zy6yx3 {
        padding-top: 2rem !important;
    }
    section.main .block-container > div:first-child {
        margin-top: 0 !important;
        padding-top: 0 !important;
    }
    section.main div[data-testid="element-container"]:has(.pap-page-header) {
        margin-top: 0 !important;
        padding-top: 0 !important;
    }
    section.main div[data-testid="stMarkdown"]:has(.pap-page-header) {
        margin-top: 0 !important;
        padding-top: 0 !important;
    }
    section.main div[data-testid="stMarkdown"]:has(.pap-page-header) h1 {
        margin-top: 0 !important;
        padding-top: 0 !important;
    }
    .pap-page-header {
        margin-top: 0 !important;
        padding-top: 0 !important;
    }
    .pap-page-header h1 {
        font-size: 1.85rem;
        font-weight: 800;
        color: #1e3a5f;
        margin: 0 0 0.75rem 0;
        padding: 0;
    }
    .pap-page-header h2 {
        font-size: 1.25rem;
        font-weight: 700;
        color: #334155;
        margin: 0.5rem 0 1rem 0;
        padding: 0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

storage = PapStorage()
auth_storage = AuthStorage()

_AUTH_QUERY_PARAM = "pap_auth"

_SUBTITULOS_PAGINA = {
    "Resumo": "Resumo da turma",
    "Relatórios": "Relatórios",
    "Definições": "Definições",
}


def _renderizar_titulo_plataforma() -> None:
    st.markdown(
        f'<div class="pap-page-header">'
        f"<h1>{html.escape(titulo_app())}</h1>"
        f"</div>",
        unsafe_allow_html=True,
    )


def _renderizar_subtitulo_pagina(pagina: str) -> None:
    subtitulo = _SUBTITULOS_PAGINA.get(pagina, pagina)
    st.markdown(
        f'<div class="pap-page-header">'
        f"<h2>{html.escape(subtitulo)}</h2>"
        f"</div>",
        unsafe_allow_html=True,
    )


def _sessao_auth() -> dict | None:
    return st.session_state.get("auth")


def _iniciar_sessao_auth(user: Utilizador, token: str) -> None:
    st.session_state["auth"] = {
        "id": user.id,
        "username": user.username,
        "nome": user.nome,
        "role": user.role,
    }
    st.session_state["auth_token"] = token


def _terminar_sessao_auth() -> None:
    token = st.session_state.get("auth_token")
    if token:
        auth_storage.revogar_sessao(str(token))
    st.session_state.pop("auth", None)
    st.session_state.pop("auth_token", None)
    _limpar_token_browser()
    if _AUTH_QUERY_PARAM in st.query_params:
        del st.query_params[_AUTH_QUERY_PARAM]


def _sincronizar_token_browser(token: str) -> None:
    tok = json.dumps(token)
    components.html(
        f"""
        <script>
        (function() {{
            var t = {tok};
            sessionStorage.setItem('pap_auth', t);
            document.cookie = 'pap_auth=' + encodeURIComponent(t) + '; path=/; SameSite=Lax';
        }})();
        </script>
        """,
        height=0,
    )


def _limpar_token_browser() -> None:
    components.html(
        """
        <script>
        sessionStorage.removeItem('pap_auth');
        document.cookie = 'pap_auth=; path=/; Max-Age=0; SameSite=Lax';
        </script>
        """,
        height=0,
    )


def _inject_auth_redirect_browser() -> None:
    components.html(
        f"""
        <script>
        (function() {{
            function gc(n) {{
                var m = document.cookie.match(new RegExp('(^| )' + n + '=([^;]+)'));
                return m ? decodeURIComponent(m[2]) : '';
            }}
            var t = gc('pap_auth') || sessionStorage.getItem('pap_auth') || '';
            if (!t) return;
            var u = new URL(window.parent.location.href);
            if (u.searchParams.get('{_AUTH_QUERY_PARAM}') === t) return;
            u.searchParams.set('{_AUTH_QUERY_PARAM}', t);
            window.parent.location.replace(u.toString());
        }})();
        </script>
        """,
        height=0,
    )


def _tentar_restaurar_sessao_auth() -> None:
    if _sessao_auth():
        return
    token = st.session_state.get("auth_token") or st.query_params.get(_AUTH_QUERY_PARAM)
    if token:
        user = auth_storage.utilizador_por_token_sessao(str(token))
        if user:
            _iniciar_sessao_auth(user, str(token))
            if _AUTH_QUERY_PARAM in st.query_params:
                del st.query_params[_AUTH_QUERY_PARAM]
            return
        st.session_state.pop("auth_token", None)
        if _AUTH_QUERY_PARAM in st.query_params:
            del st.query_params[_AUTH_QUERY_PARAM]
        _limpar_token_browser()
        return
    _inject_auth_redirect_browser()


def _pagina_login() -> None:
    titulo = titulo_app()
    st.markdown(
        f"<h1 style='text-align: center; margin-bottom: 1.5rem;'>{html.escape(titulo)}</h1>",
        unsafe_allow_html=True,
    )
    _, col, _ = st.columns([4, 2, 4])
    with col:
        st.subheader("Início de sessão")
        with st.form("login_form", enter_to_submit=True):
            username = st.text_input("Utilizador")
            password = st.text_input("Palavra-passe", type="password")
            if st.form_submit_button("Entrar", type="primary", use_container_width=True):
                user = auth_storage.autenticar(username, password)
                if user:
                    token = auth_storage.criar_sessao(user.id)
                    _iniciar_sessao_auth(user, token)
                    _sincronizar_token_browser(token)
                    st.rerun()
                else:
                    st.error("Utilizador ou palavra-passe inválidos.")


def _actualizar_nome_sessao(user_id: int, nome: str) -> None:
    sessao = _sessao_auth()
    if sessao and sessao["id"] == user_id:
        sessao["nome"] = nome


def _pagina_minha_conta(sessao: dict) -> None:
    with st.expander("A minha conta"):
        novo_nome = st.text_input("Nome", sessao["nome"], key="meu_nome")
        if st.button("Guardar nome", key="guardar_meu_nome"):
            try:
                auth_storage.alterar_nome(sessao["id"], novo_nome)
                _actualizar_nome_sessao(sessao["id"], novo_nome.strip())
                st.success("Nome actualizado.")
                st.rerun()
            except ValueError as exc:
                st.error(str(exc))
        st.divider()
        with st.form("alterar_password", enter_to_submit=False):
            atual = st.text_input("Palavra-passe actual", type="password")
            nova = st.text_input("Nova palavra-passe", type="password")
            confirmar = st.text_input("Confirmar nova palavra-passe", type="password")
            if st.form_submit_button("Alterar palavra-passe"):
                if not auth_storage.autenticar(sessao["username"], atual):
                    st.error("Palavra-passe actual incorrecta.")
                elif len(nova) < 6:
                    st.error("A nova palavra-passe deve ter pelo menos 6 caracteres.")
                elif nova != confirmar:
                    st.error("As palavras-passe não coincidem.")
                else:
                    auth_storage.alterar_password(sessao["id"], nova)
                    st.success("Palavra-passe alterada.")


def _pagina_gestao_utilizadores(sessao: dict) -> None:
    if sessao["role"] != "admin":
        return
    st.divider()
    st.header("Utilizadores")
    utilizadores = auth_storage.listar_utilizadores()

    for u in utilizadores:
        if u.id == sessao["id"]:
            continue
        estado = "activo" if u.ativo else "inactivo"
        perfil = "Administrador" if u.role == "admin" else "Professor"
        with st.expander(f"{u.nome} (@{u.username}) — {perfil}, {estado}"):
            nome_edit = st.text_input("Nome", u.nome, key=f"nome_user_{u.id}")
            c1, c2 = st.columns(2)
            if c1.button("Guardar nome", key=f"save_nome_{u.id}", use_container_width=True):
                try:
                    auth_storage.alterar_nome(u.id, nome_edit)
                    st.success("Nome actualizado.")
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))
            if c2.button(
                "Desactivar" if u.ativo else "Activar",
                key=f"toggle_{u.id}",
                use_container_width=True,
            ):
                auth_storage.definir_ativo(u.id, not u.ativo)
                st.rerun()
            with st.form(f"reset_pwd_{u.id}", enter_to_submit=False):
                pwd = st.text_input("Nova palavra-passe", type="password", key=f"pwd_{u.id}")
                if st.form_submit_button("Redefinir palavra-passe", use_container_width=True):
                    if len(pwd) < 6:
                        st.error("Mínimo 6 caracteres.")
                    else:
                        auth_storage.redefinir_password(u.id, pwd)
                        st.success("Palavra-passe redefinida.")

    st.subheader("Novo utilizador")
    with st.form("novo_utilizador", enter_to_submit=False):
        username = st.text_input("Utilizador (login)")
        nome = st.text_input("Nome")
        password = st.text_input("Palavra-passe inicial", type="password")
        role = st.selectbox("Perfil", ["professor", "admin"], format_func=lambda r: "Professor" if r == "professor" else "Administrador")
        if st.form_submit_button("Criar utilizador", type="primary"):
            if len(username.strip()) < 2:
                st.error("Indique um nome de utilizador.")
            elif len(password) < 6:
                st.error("A palavra-passe deve ter pelo menos 6 caracteres.")
            else:
                try:
                    auth_storage.criar_utilizador(username, password, nome, role)
                    st.success(f"Utilizador «{username.strip()}» criado.")
                    st.rerun()
                except ValueError as exc:
                    st.error(str(exc))



def _ia_status() -> tuple[bool, str]:
    """Estado da IA com cache curto quando indisponível (Ollama pode arrancar depois)."""
    import time

    agora = time.time()
    cache = st.session_state.get("_ia_status")
    if cache:
        ttl = 60 if cache["disp"] else 8
        if agora - cache["ts"] < ttl:
            return cache["disp"], cache["motor"]
    disp, motor = ia_disponivel()
    st.session_state["_ia_status"] = {"disp": disp, "motor": motor, "ts": agora}
    return disp, motor


def _aviso_ia_indisponivel() -> None:
    if EM_STREAMLIT_CLOUD:
        st.warning(
            "Motor IA indisponível online. Configure **Settings → Secrets** no Streamlit Cloud "
            "com `LLM_PROVIDER = \"gemini\"` e `GEMINI_API_KEY` (Ollama não funciona na cloud)."
        )
        return
    st.warning(
        "Motor IA indisponível. Opções: **Ollama** (local, gratuito — ollama.com), "
        "**ChatGPT** ou **Gemini** (chaves no `.env`). "
        "Com várias configuradas, use `LLM_PROVIDER=auto` para escolher a primeira disponível."
    )


def _inicializar_instrucoes() -> None:
    if "instrucoes" not in st.session_state:
        instr = carregar_instrucoes()
        if not instr.instrucoes_gerais and not any(instr.areas.values()):
            instr = instrucoes_default()
            guardar_instrucoes(instr)
        st.session_state.instrucoes = instr


def _gerar_pdf_seguro(docx: Path) -> tuple[Path | None, str | None]:
    resultado = gerar_pdf(docx)
    if isinstance(resultado, tuple):
        return resultado
    return resultado, None


def _apagar_ficheiros_aluno(aluno: AlunoRelatorio) -> list[str]:
    avisos: list[str] = []
    for p in (RELATORIOS_DIR / aluno.ficheiro, caminho_pdf_para_docx(aluno.ficheiro)):
        if not p.exists():
            continue
        try:
            p.unlink()
        except OSError as exc:
            avisos.append(f"Não foi possível apagar {p.name}: {exc}")
    return avisos


def _importar_ficheiros(ficheiros) -> int:
    importados = 0
    for ficheiro in ficheiros:
        destino = RELATORIOS_DIR / ficheiro.name
        destino.write_bytes(ficheiro.getbuffer())
        try:
            info = analisar_relatorio(destino)
        except Exception as exc:
            st.sidebar.error(f"Erro em {ficheiro.name}: {exc}")
            continue
        nome = nome_por_ficheiro(ficheiro.name) or info["nome"]
        storage.guardar_aluno(
            AlunoRelatorio(
                nome=nome,
                titulo_pap=info["titulo_pap"],
                tema_pap=tema_para_nome(nome) or info["titulo_pap"],
                area_pap=info["area_pap"],
                ficheiro=ficheiro.name,
                texto_extraido=info["texto_extraido"],
                importado_em=datetime.now(),
            )
        )
        importados += 1
    return importados


def _reanalisar_capa(aluno: AlunoRelatorio) -> None:
    docx = RELATORIOS_DIR / aluno.ficheiro
    if not docx.exists():
        st.error("Ficheiro do relatório não encontrado.")
        return
    info = analisar_relatorio(docx)
    storage.guardar_aluno(
        AlunoRelatorio(
            id=aluno.id,
            nome=info["nome"],
            titulo_pap=info["titulo_pap"],
            tema_pap=aluno.tema_pap,
            area_pap=info["area_pap"],
            ficheiro=aluno.ficheiro,
            texto_extraido=info["texto_extraido"],
            importado_em=aluno.importado_em,
        )
    )
    st.toast("Capa reanalisada.")
    st.rerun()


def _mostrar_pdf(aluno: AlunoRelatorio) -> None:
    if not st.toggle("Mostrar pré-visualização do relatório", key=f"showpdf_{aluno.id}"):
        st.caption("Ativa para ver o relatório (PDF ou HTML).")
        return

    docx = RELATORIOS_DIR / aluno.ficheiro
    if not docx.exists():
        st.warning(
            "Ficheiro .docx não encontrado. Importe o relatório na barra lateral "
            "ou restaure um **backup completo** (o essencial não inclui .docx)."
        )
        return

    pdf = caminho_pdf_para_docx(aluno.ficheiro)
    erro_pdf: str | None = None

    if EM_STREAMLIT_CLOUD:
        # Na cloud não há Word/LibreOffice — só PDF já existente ou HTML.
        if pdf.exists():
            pdf_bytes = pdf.read_bytes()
            try:
                from streamlit_pdf_viewer import pdf_viewer

                pdf_viewer(pdf_bytes, height=700, zoom_level=1.5, key=f"pdfview_{aluno.id}")
            except Exception:
                st.info("Use o botão abaixo para descarregar o PDF.")
            st.download_button(
                "Descarregar PDF",
                data=pdf_bytes,
                file_name=pdf.name,
                mime="application/pdf",
                key=f"dl_pdf_{aluno.id}",
            )
            return
    elif not pdf.exists():
        pdf, erro_pdf = _gerar_pdf_seguro(docx)

    if not EM_STREAMLIT_CLOUD and pdf and pdf.exists():
        pdf_bytes = pdf.read_bytes()
        try:
            from streamlit_pdf_viewer import pdf_viewer

            pdf_viewer(pdf_bytes, height=700, zoom_level=1.5, key=f"pdfview_{aluno.id}")
        except Exception:
            st.info(
                "Pré-visualização inline indisponível. "
                "Use o botão abaixo para descarregar o PDF."
            )
        st.download_button(
            "Descarregar PDF",
            data=pdf_bytes,
            file_name=pdf.name,
            mime="application/pdf",
            key=f"dl_pdf_{aluno.id}",
        )
        return

    if erro_pdf and not EM_STREAMLIT_CLOUD:
        st.warning(erro_pdf)

    html_doc, erro_html = docx_para_html(docx)
    if html_doc:
        legenda = (
            "Pré-visualização do relatório (HTML)."
            if EM_STREAMLIT_CLOUD
            else "Pré-visualização HTML (alternativa quando PDF não está disponível)."
        )
        st.caption(legenda)
        st.markdown(
            f'<div style="max-height:700px;overflow:auto;border:1px solid #e2e8f0;'
            f'padding:1rem;border-radius:8px;background:#fff">{html_doc}</div>',
            unsafe_allow_html=True,
        )
    elif erro_html and not erro_pdf:
        st.warning(erro_html)

    st.download_button(
        "Descarregar DOCX",
        data=docx.read_bytes(),
        file_name=docx.name,
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        key=f"dl_docx_{aluno.id}",
    )

    if EM_STREAMLIT_CLOUD:
        st.caption(
            "Online mostra-se em HTML. Para PDF com formatação exacta, use a app no **PC**."
        )
    elif not html_doc:
        st.caption(
            "Instale **Microsoft Word** ou **LibreOffice** no PC para gerar PDF. "
            f"O .docx está em `{docx}`."
        )


def _guardar_criterio_cb(aluno_id: int, criterio: CriterioAvaliacao) -> None:
    """Guarda automaticamente ao alterar a nota ou o comentario (sem botao)."""
    nota = st.session_state.get(f"nota_{criterio.value}_{aluno_id}")
    comentario = st.session_state.get(f"com_{criterio.value}_{aluno_id}", "")
    if nota is None:
        return
    existente = storage.obter_avaliacoes(aluno_id).get(criterio)
    if criterio in CRITERIOS_MANUAIS:
        fonte = "manual"
    elif existente:
        fonte = existente.fonte if "editado" in existente.fonte else f"{existente.fonte} (editado)"
    else:
        fonte = "manual"
    storage.guardar_avaliacao(
        aluno_id,
        ResultadoCriterio(
            criterio=criterio, nota=int(nota), comentario=comentario, fonte=fonte
        ),
    )


def _renderizar_criterio_editavel(
    aluno: AlunoRelatorio, criterio: CriterioAvaliacao, avaliacoes: dict
) -> None:
    existente = avaliacoes.get(criterio)
    nota_default = int(existente.nota) if existente else 10
    comentario_default = existente.comentario if existente else ""
    with st.container(border=True):
        st.markdown(f"**{CRITERIO_LABELS[criterio]}**")
        col_nota, col_com = st.columns([1, 5])
        with col_nota:
            st.number_input(
                f"Nota (1-{NOTA_MAXIMA})",
                min_value=1,
                max_value=NOTA_MAXIMA,
                value=nota_default,
                step=1,
                key=f"nota_{criterio.value}_{aluno.id}",
                on_change=_guardar_criterio_cb,
                args=(aluno.id, criterio),
            )
        with col_com:
            st.text_area(
                "Comentário",
                comentario_default,
                height=110,
                key=f"com_{criterio.value}_{aluno.id}",
                on_change=_guardar_criterio_cb,
                args=(aluno.id, criterio),
            )


def _renderizar_avaliacoes_editaveis(aluno: AlunoRelatorio, avaliacoes: dict) -> None:
    secao = SecaoAvaliacao.RELATORIO
    pct = int(SECAO_PONDERACAO[secao] * 100)
    st.markdown(f"**{SECAO_LABELS[secao]}** ({pct}%)")
    st.caption("Critérios avaliados pela IA a partir do relatório (excepto o marcado como manual).")
    for criterio in CRITERIOS_POR_SECAO[secao]:
        if criterio == CRITERIO_MANUAL:
            st.caption("Critério manual (professor)")
        _renderizar_criterio_editavel(aluno, criterio, avaliacoes)


def _nota_atual(aluno: AlunoRelatorio, criterio: CriterioAvaliacao, avaliacoes: dict):
    """Nota atual do critério: valor do campo (se editado) ou o guardado."""
    key = f"nota_{criterio.value}_{aluno.id}"
    if key in st.session_state:
        return st.session_state[key]
    if criterio in avaliacoes:
        return avaliacoes[criterio].nota
    return None


def _media_atual(aluno: AlunoRelatorio, avaliacoes: dict):
    return avaliacao_final(avaliacoes)


def _estrutura_do_texto(texto: str) -> list[str]:
    """Le a lista de titulos guardada no inicio do texto_extraido."""
    if "ESTRUTURA DO RELATORIO" not in texto or "CONTEUDO:" not in texto:
        return []
    bloco = texto.split("CONTEUDO:", 1)[0]
    linhas = bloco.splitlines()[1:]  # ignora a linha de cabecalho
    return [l.strip() for l in linhas if l.strip()]


def _agrupar_capitulos(titulos: list[str]) -> list[dict]:
    """Agrupa os titulos numerados por capitulo (nivel 1) com os subtopicos (nivel 2)."""
    chaves = list(CAPITULOS_PAP.keys())
    capitulos: list[dict] = []
    atual: dict | None = None
    for t in titulos:
        m = re.match(r"^(\d+)(?:\.(\d+))?(?:\.(\d+))?\.?\s+(.+)$", t)
        if not m:
            continue
        n1, n2, n3, _titulo = m.groups()
        if n2 is None:  # capitulo principal (1., 2., ...)
            num = int(n1)
            chave = chaves[num - 1] if 1 <= num <= len(chaves) else None
            atual = {"titulo": t, "chave": chave, "subs": []}
            capitulos.append(atual)
        elif n3 is None and atual is not None:  # subtopico principal (x.y)
            atual["subs"].append(t)
    return capitulos


def _renderizar_resumo_capitulos(aluno: AlunoRelatorio) -> None:
    capitulos = _agrupar_capitulos(_estrutura_do_texto(aluno.texto_extraido))
    if not capitulos:
        st.caption(
            "Sem estrutura detetada neste relatório. "
            "Usa **Re-analisar capa** para reprocessar o ficheiro."
        )
        return

    resumos_ia = storage.obter_resumos_capitulos(aluno.id)
    disp, motor = _ia_status()

    if disp:
        rotulo = "Atualizar resumo com IA" if resumos_ia else "Resumir relatório com IA"
        if st.button(rotulo, type="primary", key=f"resumocap_{aluno.id}"):
            barra = st.progress(0.0, text="A iniciar...")
            try:
                def _prog(feitos: int, total: int, titulo: str) -> None:
                    barra.progress(
                        (feitos + 1) / max(total, 1),
                        text=f"A resumir ({feitos + 1}/{total}): {titulo[:45]}…",
                    )

                novos = resumir_capitulos(aluno.texto_extraido, on_progress=_prog)
                storage.guardar_resumos_capitulos(aluno.id, novos)
                barra.empty()
                st.rerun()
            except Exception as exc:
                barra.empty()
                st.error(str(exc))
    else:
        _aviso_ia_indisponivel()

    if not resumos_ia:
        st.info("Carrega em **Resumir relatório com IA** para gerar o resumo de cada subcapítulo.")
        return

    for cap in capitulos:
        chave = cap.get("chave")
        resumo = resumos_ia.get(chave) if chave else None
        if not resumo:
            continue
        with st.container(border=True):
            st.markdown(f"**{cap['titulo']}**")
            st.markdown(resumo)


def _metrica(rotulo: str, valor: str, alinhar: str = "center") -> None:
    st.markdown(
        f"<div style='text-align:{alinhar}'>"
        f"<div style='color:#808495;font-size:0.8rem'>{rotulo}</div>"
        f"<div style='font-size:1.6rem;font-weight:600'>{valor}</div>"
        f"</div>",
        unsafe_allow_html=True,
    )


def _renderizar_tab_aluno(aluno: AlunoRelatorio) -> None:
    avaliacoes = storage.obter_avaliacoes(aluno.id)
    col1, col2, col3 = st.columns([2, 3, 1])
    with col1:
        novo_nome = st.text_input("Nome do aluno", aluno.nome, key=f"nome_{aluno.id}")
    with col2:
        novo_titulo = st.text_input(
            "Título da PAP", aluno.titulo_pap, key=f"titulo_{aluno.id}"
        )
    with col3:
        areas = list(AreaPAP)
        idx = areas.index(aluno.area_pap) if aluno.area_pap in areas else 0
        nova_area = st.selectbox(
            "Área da PAP", areas, index=idx,
            format_func=lambda a: AREA_LABELS[a], key=f"area_{aluno.id}",
        )
    novo_tema = st.text_input("Tema da PAP", aluno.tema_pap, key=f"tema_{aluno.id}")

    c1, c2, c3 = st.columns(3)
    if c1.button("Atualizar dados", key=f"upd_{aluno.id}"):
        storage.atualizar_aluno(
            aluno.id,
            nome=novo_nome,
            titulo_pap=novo_titulo,
            tema_pap=novo_tema,
            area_pap=nova_area,
        )
        st.rerun()
    if c2.button("Re-analisar capa", key=f"reparse_{aluno.id}"):
        _reanalisar_capa(aluno)
    if c3.button("Remover aluno", key=f"del_{aluno.id}"):
        avisos = _apagar_ficheiros_aluno(aluno)
        storage.remover_aluno(aluno.id)
        for aviso in avisos:
            st.warning(aviso)
        st.toast(f"{aluno.nome} removido.")
        st.rerun()

    st.divider()
    col_a, col_b, col_c = st.columns([4, 4, 4], vertical_alignment="center")
    with col_a:
        _metrica("Relatório", aluno.nome, alinhar="left")
    with col_b:
        media_rel = media_secao_arredondada(avaliacoes, SecaoAvaliacao.RELATORIO)
        _metrica(
            "Nota relatório",
            f"{media_rel}/{NOTA_MAXIMA}" if media_rel is not None else "-",
        )
    with col_c:
        st.markdown(
            "<div style='text-align:center;color:#808495;font-size:0.8rem'>Avaliação automática</div>",
            unsafe_allow_html=True,
        )
        disp, motor = _ia_status()
        _, cbtn, _ = st.columns([1, 2, 1])
        if not disp:
            _aviso_ia_indisponivel()
        elif cbtn.button("Avaliar relatório", type="primary", key=f"avaliar_{aluno.id}", use_container_width=True):
            with st.spinner("A avaliar... pode demorar 1-2 minutos com Ollama."):
                try:
                    resultados = avaliar_relatorio(
                        novo_nome, novo_tema or novo_titulo, nova_area.value,
                        aluno.texto_extraido, carregar_instrucoes(),
                    )
                    for r in resultados:
                        storage.guardar_avaliacao(aluno.id, r)
                        st.session_state[f"nota_{r.criterio.value}_{aluno.id}"] = int(r.nota)
                        st.session_state[f"com_{r.criterio.value}_{aluno.id}"] = r.comentario
                    st.success("Avaliação concluída!")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))

    st.divider()
    st.subheader("Avaliação do relatório")
    avaliacoes = storage.obter_avaliacoes(aluno.id)
    _renderizar_avaliacoes_editaveis(aluno, avaliacoes)

    st.divider()
    st.subheader("Resumo por capítulo")
    st.caption("Resumo com IA do conteúdo abordado em cada subcapítulo (não copia o texto do aluno).")
    _renderizar_resumo_capitulos(aluno)

    st.divider()
    st.subheader("Pré-visualização (PDF)")
    _mostrar_pdf(aluno)


def _pagina_configuracao_app(sessao: dict) -> None:
    if sessao["role"] != "admin":
        return
    st.header("Aplicação")
    config = carregar_configuracao_app()
    titulo = st.text_input("Título da aplicação", config.titulo)
    c1, _ = st.columns(2)
    if c1.button("Guardar título", type="primary", key="guardar_titulo_app"):
        guardar_configuracao_app(ConfiguracaoApp(titulo=titulo.strip() or TITULO_PADRAO))
        st.success("Título guardado.")
        st.rerun()
    st.divider()


def _pagina_nomes_alunos() -> None:
    st.header("Nomes da turma")
    st.caption(
        "Defina o **nome completo** de cada aluno e a **chave no ficheiro** "
        "(texto que aparece no nome do .docx importado, ex.: `AdrianoSalucombo`). "
        "Use **Aplicar aos alunos importados** para corrigir nomes já na base de dados."
    )
    entries = carregar_nomes_turma()
    df = pd.DataFrame(
        [
            {
                "Nome completo": e.nome,
                "Chave no ficheiro": e.chave_ficheiro,
                "Tema PAP": e.tema,
            }
            for e in entries
        ]
    )
    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Nome completo": st.column_config.TextColumn(required=True),
            "Chave no ficheiro": st.column_config.TextColumn(
                help="Parte do nome do ficheiro .docx (ex.: VladimiroPankratau)",
            ),
            "Tema PAP": st.column_config.TextColumn(
                help="Tema oficial da PAP (opcional; usado na grelha e exportação)",
            ),
        },
        key="editor_nomes_turma",
    )
    c1, c2, c3 = st.columns(3)
    if c1.button("Guardar nomes", type="primary", key="guardar_nomes_turma"):
        novos = [
            AlunoTurma(
                nome=str(row["Nome completo"]).strip(),
                chave_ficheiro=str(row.get("Chave no ficheiro") or "").strip(),
                tema=str(row.get("Tema PAP") or "").strip(),
            )
            for _, row in edited.iterrows()
            if str(row.get("Nome completo", "")).strip()
        ]
        if not novos:
            st.error("Indique pelo menos um aluno.")
        else:
            guardar_nomes_turma(novos)
            st.success("Nomes guardados.")
            st.rerun()
    if c2.button("Aplicar aos alunos importados", key="aplicar_nomes_turma"):
        novos = [
            AlunoTurma(
                nome=str(row["Nome completo"]).strip(),
                chave_ficheiro=str(row.get("Chave no ficheiro") or "").strip(),
                tema=str(row.get("Tema PAP") or "").strip(),
            )
            for _, row in edited.iterrows()
            if str(row.get("Nome completo", "")).strip()
        ]
        if novos:
            guardar_nomes_turma(novos)
        total, avisos = aplicar_nomes_a_alunos(storage)
        if total:
            st.success(f"{total} aluno(s) actualizado(s).")
        else:
            st.info("Nenhum aluno precisou de actualização.")
        for aviso in avisos[:10]:
            st.warning(aviso)
        if len(avisos) > 10:
            st.caption(f"… e mais {len(avisos) - 10} avisos.")
        st.rerun()
    if c3.button("Restaurar lista padrão", key="reset_nomes_turma"):
        guardar_nomes_turma(nomes_turma_default())
        st.success("Lista padrão restaurada.")
        st.rerun()
    st.divider()


def _pagina_backup_dados(sessao: dict) -> None:
    if sessao["role"] != "admin":
        return
    st.header("Dados e sincronização")
    st.caption(
        "O **PC local** e o **Streamlit Cloud** são instalações separadas — não partilham "
        "dados automaticamente. Use backup para copiar alunos, notas e relatórios entre os dois."
    )
    st.info(f"Armazenamento actual: **{descricao_armazenamento()}**")

    col_dl, col_up = st.columns(2)
    with col_dl:
        try:
            zip_essencial = exportar_backup(completo=False)
            mb_ess = len(zip_essencial) / (1024 * 1024)
            st.download_button(
                f"⬇ Backup essencial ({mb_ess:.1f} MB)",
                data=zip_essencial,
                file_name=f"avaliapap_essencial_{datetime.now():%Y%m%d_%H%M}.zip",
                mime="application/zip",
                use_container_width=True,
                key="dl_backup_essencial",
                help="Base de dados, notas, configuração — ideal para PC ↔ Cloud (limite 200 MB).",
            )
            zip_completo = exportar_backup(completo=True)
            mb_comp = len(zip_completo) / (1024 * 1024)
            st.download_button(
                f"⬇ Backup completo ({mb_comp:.0f} MB)",
                data=zip_completo,
                file_name=f"avaliapap_completo_{datetime.now():%Y%m%d_%H%M}.zip",
                mime="application/zip",
                use_container_width=True,
                key="dl_backup_completo",
                help="Inclui ficheiros .docx/.pdf — só recomendado no PC local.",
            )
        except Exception as exc:
            st.error(f"Erro ao exportar: {exc}")
    with col_up:
        upload = st.file_uploader(
            "Importar backup (.zip)",
            type=["zip"],
            key="upload_backup_dados",
        )
        if upload:
            mb_up = upload.size / (1024 * 1024) if upload.size else 0
            st.caption(f"Ficheiro seleccionado: {upload.name} ({mb_up:.1f} MB)")
        if upload and st.button("Restaurar backup", type="primary", key="btn_import_backup"):
            try:
                n, avisos = importar_backup(upload.getvalue())
                st.success(f"Backup restaurado ({n} ficheiros). A recarregar…")
                for aviso in avisos:
                    st.warning(aviso)
                st.rerun()
            except Exception as exc:
                st.error(str(exc))

    st.caption(
        "**Essencial** — alunos, notas, textos para IA (cabe online). "
        "**Completo** — inclui .docx (só PC; na cloud use essencial + importe .docx na sidebar). "
        "Limite upload Streamlit Cloud: **200 MB**."
    )
    st.divider()


def _pagina_configuracao() -> None:
    sessao = _sessao_auth()
    if sessao:
        _pagina_configuracao_app(sessao)

    _pagina_nomes_alunos()

    sessao = _sessao_auth()
    if sessao:
        _pagina_backup_dados(sessao)

    st.header("Instruções de avaliação")
    instr = carregar_instrucoes()
    geral = st.text_area("Instruções gerais", instr.instrucoes_gerais, height=150)

    st.subheader("Por capítulo do relatório")
    st.caption("Estrutura comum a todos os relatórios de PAP. Define o que se espera em cada capítulo.")
    capitulos_edit = {}
    areas_edit = dict(instr.areas)
    for chave, titulo in CAPITULOS_PAP.items():
        capitulos_edit[chave] = st.text_area(
            titulo, instr.capitulos.get(chave, ""), height=120, key=f"cap_{chave}"
        )
        # O capítulo da Implementação é específico de cada área da PAP
        if chave == "implementacao":
            st.caption(
                "↳ Instruções específicas por área da PAP "
                "(o capítulo «Implementação do projeto» depende da área)."
            )
            areas = ["website", "aplicacao_movel", "robotica", "jogo"]
            for area, aba in zip(areas, st.tabs([AREA_LABELS[AreaPAP(a)] for a in areas])):
                with aba:
                    areas_edit[area] = st.text_area(
                        f"Implementação — {AREA_LABELS[AreaPAP(area)]}",
                        instr.areas.get(area, ""),
                        height=140,
                        key=f"inst_{area}",
                        label_visibility="collapsed",
                    )

    c1, c2 = st.columns(2)
    if c1.button("Guardar", type="primary"):
        from app.models import InstrucoesAvaliacao
        guardar_instrucoes(
            InstrucoesAvaliacao(
                instrucoes_gerais=geral, areas=areas_edit, capitulos=capitulos_edit
            )
        )
        st.success("Guardado.")
    if c2.button("Restaurar padrão"):
        guardar_instrucoes(instrucoes_default())
        st.rerun()

    sessao = _sessao_auth()
    if sessao:
        _pagina_minha_conta(sessao)
        _pagina_gestao_utilizadores(sessao)


def _avaliar_todos(alunos: list[AlunoRelatorio]) -> None:
    instr = carregar_instrucoes()
    barra = st.progress(0.0, text="A avaliar...")
    erros: list[str] = []
    total = len(alunos)
    for i, aluno in enumerate(alunos, start=1):
        barra.progress(i / total, text=f"A avaliar {aluno.nome} ({i}/{total})...")
        try:
            resultados = avaliar_relatorio(
                aluno.nome,
                aluno.tema_pap or aluno.titulo_pap,
                aluno.area_pap.value,
                aluno.texto_extraido,
                instr,
            )
            for r in resultados:
                storage.guardar_avaliacao(aluno.id, r)
                st.session_state[f"nota_{r.criterio.value}_{aluno.id}"] = int(r.nota)
                st.session_state[f"com_{r.criterio.value}_{aluno.id}"] = r.comentario
        except Exception as exc:
            erros.append(f"{aluno.nome}: {exc}")
    barra.empty()
    if erros:
        st.warning("Alguns relatórios falharam:\n\n- " + "\n- ".join(erros))
    else:
        st.success("Todos os relatórios foram avaliados!")


LARGURA_CRITERIO = 340
ALTURA_LINHA_GRELHA = 38
ALTURA_CABECALHO_GRELHA = 44
PAP_GRID_LAYOUT_V = 3
LARGURA_MIN_ALUNO = 40
COR_CELULA_MUITO_BOM = "#dcfce7"
COR_CELULA_BOM = "#dbeafe"
COR_CELULA_SUFICIENTE = "#fef9c3"
COR_CELULA_INSUFICIENTE = "#fee2e2"
COR_TEXTO_MUITO_BOM = "#15803d"
COR_TEXTO_BOM = "#1d4ed8"
COR_TEXTO_SUFICIENTE = "#a16207"
COR_TEXTO_INSUFICIENTE = "#b91c1c"

CRITERIO_ICONES: dict[CriterioAvaliacao, str] = {
    CriterioAvaliacao.AUTONOMIA: "👤",
    CriterioAvaliacao.FASES_EXECUCAO: "📋",
    CriterioAvaliacao.DEFINICAO_OBJETIVOS: "🎯",
    CriterioAvaliacao.PERTINENCIA_CRIATIVIDADE: "💡",
    CriterioAvaliacao.PLANEAMENTO_RECURSOS: "📅",
    CriterioAvaliacao.QUALIDADE_CIENTIFICA: "🔬",
    CriterioAvaliacao.RESPONSABILIDADE_PROJETO: "⏱",
    CriterioAvaliacao.OBJETIVIDADE: "📝",
    CriterioAvaliacao.PERTINENCIA: "📌",
    CriterioAvaliacao.DIFICULDADES: "🛠",
    CriterioAvaliacao.ANALISE_CRITICA: "🔍",
    CriterioAvaliacao.RESPONSABILIDADE_RELATORIO: "⏱",
    CriterioAvaliacao.EXPRESSAO_ORAL: "🎤",
    CriterioAvaliacao.CAPACIDADE_SINTESE: "📊",
    CriterioAvaliacao.RECURSOS_APRESENTACAO: "🖥",
    CriterioAvaliacao.ARGUMENTACAO_DEFESA: "💬",
    CriterioAvaliacao.AUTOAVALIACAO: "✍",
}

SECAO_ICONES = {
    SecaoAvaliacao.DESENVOLVIMENTO: "📋",
    SecaoAvaliacao.RELATORIO: "📗",
    SecaoAvaliacao.APRESENTACAO: "🎤",
}

_CSS_AGGRID_RESUMO = {
    ".ag-root-wrapper": {
        "width": "100% !important",
        "overflow": "hidden !important",
        "border-radius": "0 0 12px 12px !important",
        "border": "1px solid #e2e8f0 !important",
        "border-top": "none !important",
        "box-shadow": "0 1px 3px rgba(15, 23, 42, 0.05)",
    },
    ".ag-root-wrapper-body": {
        "width": "100% !important",
    },
    ".ag-theme-balham": {
        "--ag-borders": "none",
        "--ag-border-color": "#e2e8f0",
        "--ag-row-border-width": "0px",
        "--ag-row-border-color": "transparent",
        "--ag-cell-horizontal-border": "none",
        "--ag-header-column-separator-display": "block",
        "--ag-header-column-separator-height": "100%",
        "--ag-header-column-separator-color": "#e2e8f0",
        "font-family": "inherit",
    },
    ".ag-header": {
        "background-color": "#ffffff !important",
        "border-bottom": "1px solid #e2e8f0 !important",
        "height": "44px !important",
        "min-height": "44px !important",
        "max-height": "44px !important",
    },
    ".ag-header-cell": {
        "background-color": "#ffffff !important",
        "padding": "0 !important",
        "border-right": "1px solid #e2e8f0 !important",
    },
    ".ag-header-cell-label": {
        "justify-content": "center !important",
        "width": "100%",
    },
    ".ag-pinned-left-header .ag-header-cell": {
        "border-right": "1px solid #e2e8f0 !important",
    },
    ".pap-hdr-criterio": {
        "padding-left": "12px",
        "font-size": "0.78rem",
        "font-weight": "600",
        "color": "#64748b",
        "display": "flex",
        "align-items": "center",
        "height": "100%",
    },
    ".pap-hdr-pill": {
        "display": "inline-block",
        "background": "#f1f5f9",
        "border": "1px solid #e2e8f0",
        "border-radius": "8px",
        "padding": "4px 7px",
        "font-size": "0.68rem",
        "font-weight": "700",
        "color": "#334155",
        "letter-spacing": "0.03em",
        "min-width": "30px",
        "text-align": "center",
        "line-height": "1.1",
        "position": "relative",
    },
    ".pap-hdr-pill-vazia": {
        "color": "#94a3b8",
        "background": "#f8fafc",
    },
    ".pap-sort": {
        "margin-left": "4px",
        "color": "#94a3b8",
        "font-size": "0.7rem",
    },
    ".ag-cell": {
        "font-size": "13px",
        "box-sizing": "border-box !important",
        "border-right": "1px solid #e2e8f0 !important",
        "border-bottom": "1px solid #e2e8f0 !important",
        "border-top": "none !important",
        "padding": "0 2px !important",
    },
    ".ag-center-cols-container .ag-cell": {
        "text-align": "center !important",
    },
    ".ag-center-cols-container .ag-cell-value": {
        "text-align": "center !important",
    },
    ".ag-cell-value": {
        "overflow": "visible !important",
        "text-overflow": "clip !important",
    },
    ".ag-pinned-left-cols-container .ag-cell": {
        "border-right": "1px solid #e2e8f0 !important",
    },
    ".ag-row": {
        "height": "38px !important",
        "max-height": "38px !important",
        "border": "none !important",
    },
    ".ag-row-last .ag-cell": {
        "border-bottom": "none !important",
    },
    ".pap-row-total .ag-cell": {
        "font-weight": "700 !important",
    },
    ".ag-header-row": {
        "height": "44px !important",
        "max-height": "44px !important",
    },
    ".ag-body-viewport": {
        "overflow": "hidden !important",
    },
    ".ag-center-cols-viewport": {
        "overflow": "hidden !important",
    },
    ".ag-header-viewport": {
        "overflow": "hidden !important",
    },
    ".ag-body-horizontal-scroll": {
        "display": "none !important",
        "height": "0 !important",
        "min-height": "0 !important",
        "max-height": "0 !important",
    },
    ".ag-body-vertical-scroll": {
        "display": "none !important",
        "width": "0 !important",
        "min-width": "0 !important",
        "max-width": "0 !important",
    },
    ".ag-horizontal-left-spacer": {
        "display": "none !important",
    },
}

_JS_HEADER_CRITERIO = JsCode(
    """
class PapHeaderCriterio {
    init(params) {
        this.eGui = document.createElement('div');
        this.eGui.className = 'pap-hdr-criterio';
        this.eGui.innerHTML = 'Critério <span class="pap-sort">⇅</span>';
    }
    getGui() {
        return this.eGui;
    }
}
"""
)

_JS_HEADER_ALUNO = JsCode(
    """
class PapHeaderAluno {
    init(params) {
        this.eGui = document.createElement('div');
        this.eGui.style.display = 'flex';
        this.eGui.style.alignItems = 'center';
        this.eGui.style.justifyContent = 'center';
        this.eGui.style.width = '100%';
        this.eGui.style.height = '100%';
        var tip = params.column.colDef.headerTooltip || params.displayName;
        var pill = document.createElement('span');
        pill.className = 'pap-hdr-pill';
        if (params.column.colDef.editable === false) {
            pill.className += ' pap-hdr-pill-vazia';
        }
        pill.setAttribute('data-tip', tip);
        pill.textContent = params.displayName;
        this.eGui.appendChild(pill);
    }
    getGui() {
        return this.eGui;
    }
}
"""
)

_JS_ESTILO_CELULA = JsCode(
    """
function(params) {
    if (params.colDef.field === 'Critério') {
        return {
            'backgroundColor': '#ffffff',
            'fontSize': '12px',
            'textAlign': 'left',
            'paddingLeft': '10px',
            'fontWeight': '500',
            'color': '#334155',
        };
    }
    var v = params.value;
    var base = {
        'textAlign': 'center',
        'fontWeight': '700',
        'fontSize': '13px',
    };
    if (v === null || v === undefined || v === '') {
        return Object.assign(base, {
            'backgroundColor': '#ffffff',
            'color': '#94a3b8',
            'fontWeight': '600',
        });
    }
    v = Number(v);
    if (isNaN(v)) {
        return Object.assign(base, {'backgroundColor': '#ffffff'});
    }
    if (v >= 18) {
        return Object.assign(base, {
            'backgroundColor': '#dcfce7',
            'color': '#15803d',
        });
    }
    if (v >= 14) {
        return Object.assign(base, {
            'backgroundColor': '#dbeafe',
            'color': '#1d4ed8',
        });
    }
    if (v >= 10) {
        return Object.assign(base, {
            'backgroundColor': '#fef9c3',
            'color': '#a16207',
        });
    }
    return Object.assign(base, {
        'backgroundColor': '#fee2e2',
        'color': '#b91c1c',
    });
}
"""
)

_JS_FORMATO_CELULA = JsCode(
    """
function(params) {
    if (params.colDef.field === 'Critério') {
        return params.value;
    }
    if (params.value === null || params.value === undefined || params.value === '') {
        return '—';
    }
    var v = Number(params.value);
    if (isNaN(v)) {
        return params.value;
    }
    return (v % 1 === 0) ? String(Math.round(v)) : v.toFixed(1);
}
"""
)

_ROTULOS_LINHA_SO_LEITURA = frozenset({"Total:", "Média ponderada", "Nota Final"})

_JS_CELULA_EDITAVEL = JsCode(
    """
function(params) {
    var c = params.data['Critério'];
    if (c === 'Total:' || c === 'Média ponderada' || c === 'Nota Final') {
        return false;
    }
    return true;
}
"""
)


def _rotulo_criterio_resumo(criterio: CriterioAvaliacao) -> str:
    icon = CRITERIO_ICONES.get(criterio, "•")
    return f"{icon}  {CRITERIO_LABELS[criterio]}"


def _iniciais_nome(nome: str) -> str:
    partes = [p for p in nome.split() if p]
    if len(partes) >= 3:
        return "".join(p[0].upper() for p in partes[:3])
    if len(partes) == 2:
        return (partes[0][0] + partes[1][0] + (partes[1][1] if len(partes[1]) > 1 else partes[1][0])).upper()
    return nome[:3].upper()


def _classe_nota(nota: float | int | None) -> str:
    if nota is None:
        return "vazia"
    n = float(nota)
    if n >= 18:
        return "muito_bom"
    if n >= 14:
        return "bom"
    if n >= 10:
        return "suficiente"
    return "insuficiente"


def _metricas_resumo_turma(
    colunas: list[tuple[str, AlunoRelatorio | None]],
    avaliacoes: dict[int, dict],
) -> dict:
    alunos_ativos = [a for _, a in colunas if a is not None]
    n_total = len(alunos_ativos)
    notas_finais: list[int] = []
    parcial = 0
    for aluno in alunos_ativos:
        av = avaliacoes.get(aluno.id, {})
        if av:
            parcial += 1
        nf = nota_final_arredondada(av)
        if nf is not None:
            notas_finais.append(nf)
    pct = round(100 * parcial / n_total) if n_total else 0
    media = sum(notas_finais) / len(notas_finais) if notas_finais else None
    melhor = max(notas_finais) if notas_finais else None
    return {
        "total": n_total,
        "avaliados": parcial,
        "pct_avaliados": pct,
        "media": media,
        "melhor": melhor,
    }


def _html_metricas_resumo(m: dict) -> str:
    media_txt = f"{m['media']:.1f}".replace(".", ",") if m["media"] is not None else "—"
    melhor_txt = str(m["melhor"]) if m["melhor"] is not None else "—"
    return f"""
<div class="pap-metricas">
  <div class="pap-metric-card">
    <div class="pap-metric-icone pap-icone-azul">👥</div>
    <div class="pap-metric-corpo">
      <div class="pap-metric-label">Alunos</div>
      <div class="pap-metric-num">{m["total"]}</div>
      <div class="pap-metric-sub">Total de alunos</div>
    </div>
  </div>
  <div class="pap-metric-card">
    <div class="pap-metric-icone pap-icone-verde">✓</div>
    <div class="pap-metric-corpo">
      <div class="pap-metric-label">Avaliados (geral)</div>
      <div class="pap-metric-num">{m["avaliados"]}</div>
      <div class="pap-metric-sub">{m["pct_avaliados"]}% do total</div>
    </div>
  </div>
  <div class="pap-metric-card">
    <div class="pap-metric-icone pap-icone-amarelo">📊</div>
    <div class="pap-metric-corpo">
      <div class="pap-metric-label">Média da turma</div>
      <div class="pap-metric-num">{media_txt}</div>
      <div class="pap-metric-sub">Entre alunos avaliados</div>
    </div>
  </div>
  <div class="pap-metric-card">
    <div class="pap-metric-icone pap-icone-roxo">★</div>
    <div class="pap-metric-corpo">
      <div class="pap-metric-label">Melhor nota</div>
      <div class="pap-metric-num">{melhor_txt}</div>
      <div class="pap-metric-sub">Nota mais alta</div>
    </div>
  </div>
</div>
"""


def _html_legenda_notas(compacta: bool = False) -> str:
    classe = "pap-legenda pap-legenda-compacta" if compacta else "pap-legenda"
    return f"""
<span class="{classe}">
  <span class="pap-legenda-item"><i class="pap-legenda-cor pap-cor-muito-bom"></i> Muito bom (18–20)</span>
  <span class="pap-legenda-item"><i class="pap-legenda-cor pap-cor-bom"></i> Bom (14–17)</span>
  <span class="pap-legenda-item"><i class="pap-legenda-cor pap-cor-suficiente"></i> Suficiente (10–13)</span>
  <span class="pap-legenda-item"><i class="pap-legenda-cor pap-cor-insuficiente"></i> Insuficiente (0–9)</span>
  <span class="pap-legenda-item"><i class="pap-legenda-cor pap-cor-vazia"></i> Não avaliado (—)</span>
</span>
"""


def _exportar_resumo_excel(
    colunas: list[tuple[str, AlunoRelatorio | None]],
    avaliacoes: dict[int, dict],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    fills = {
        "muito_bom": PatternFill("solid", fgColor="86EFAC"),
        "bom": PatternFill("solid", fgColor="93C5FD"),
        "suficiente": PatternFill("solid", fgColor="FDE047"),
        "insuficiente": PatternFill("solid", fgColor="FCA5A5"),
        "vazia": PatternFill("solid", fgColor="E5E7EB"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.cell(1, 1, "Critério").font = Font(bold=True)
    for col_idx, (nome, aluno) in enumerate(colunas, start=2):
        cell = ws.cell(1, col_idx, _iniciais_nome(nome))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if aluno is None:
            cell.value = f"{_iniciais_nome(nome)} (s/ relatório)"
    row = 2
    chaves = [_chave_coluna_grelha(a, i) for i, (_, a) in enumerate(colunas)]
    for secao in SecaoAvaliacao:
        ws.cell(row, 1, SECAO_LABELS[secao]).font = Font(bold=True)
        row += 1
        criterios = CRITERIOS_POR_SECAO[secao]
        for criterio in criterios:
            ws.cell(row, 1, CRITERIO_LABELS[criterio])
            for col_idx, ((_, aluno), chave) in enumerate(zip(colunas, chaves), start=2):
                val = None
                if aluno is not None:
                    av = avaliacoes.get(aluno.id, {})
                    if criterio in av:
                        val = int(av[criterio].nota)
                cell = ws.cell(row, col_idx, val if val is not None else "—")
                cell.alignment = Alignment(horizontal="center")
                cell.fill = fills[_classe_nota(val)]
            row += 1
        ws.cell(row, 1, "Total:").font = Font(bold=True)
        for col_idx, ((_, aluno), chave) in enumerate(zip(colunas, chaves), start=2):
            val = None
            if aluno is not None:
                media = media_secao_arredondada(avaliacoes[aluno.id], secao)
                val = media if media is not None else None
            cell = ws.cell(row, col_idx, val if val is not None else "—")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fills[_classe_nota(val)]
        row += 1
    for rotulo, fn in (
        ("Média ponderada", lambda av: avaliacao_final(av)),
        ("Nota Final", lambda av: nota_final_arredondada(av)),
    ):
        ws.cell(row, 1, rotulo).font = Font(bold=True)
        for col_idx, (nome, aluno) in enumerate(colunas, start=2):
            val = fn(avaliacoes[aluno.id]) if aluno is not None else None
            cell = ws.cell(row, col_idx, val if val is not None else "—")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fills[_classe_nota(val)]
        row += 1
    rotulo_auto = CRITERIO_LABELS[CriterioAvaliacao.AUTOAVALIACAO]
    ws.cell(row, 1, rotulo_auto)
    for col_idx, (_, aluno) in enumerate(colunas, start=2):
        val = None
        if aluno is not None:
            av = avaliacoes.get(aluno.id, {})
            c = CriterioAvaliacao.AUTOAVALIACAO
            val = int(av[c].nota) if c in av else None
        cell = ws.cell(row, col_idx, val if val is not None else "—")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fills[_classe_nota(val)]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _larguras_colunas(n_alunos: int) -> tuple[int, int]:
    """Largura da coluna Critério (fixa) e mínimo por aluno."""
    return LARGURA_CRITERIO, LARGURA_MIN_ALUNO


_JS_LINHA_TOTAL = JsCode(
    """
function(params) {
    var c = params.data['Critério'];
    if (c === 'Total:' || c === 'Média ponderada' || c === 'Nota Final') {
        return 'pap-row-total';
    }
    return null;
}
"""
)


def _altura_grelha(n_linhas: int) -> int:
    """Altura exacta: cabeçalho + linhas, sem espaço vazio por baixo."""
    return ALTURA_CABECALHO_GRELHA + n_linhas * ALTURA_LINHA_GRELHA


def _fixar_opcoes_grelha(gb: GridOptionsBuilder) -> dict:
    """Remove auto-size que redimensiona colunas após cada edição."""
    opcoes = gb.build()
    opcoes.pop("autoSizeStrategy", None)
    opcoes["suppressAutoSize"] = True
    return opcoes


def _chave_coluna_grelha(aluno: AlunoRelatorio | None, indice: int) -> str:
    if aluno is not None:
        return f"id_{aluno.id}"
    return f"vazio_{indice}"


def _html_barra_secao(secao: SecaoAvaliacao) -> str:
    classes = {
        SecaoAvaliacao.DESENVOLVIMENTO: "pap-secao-a",
        SecaoAvaliacao.RELATORIO: "pap-secao-b",
        SecaoAvaliacao.APRESENTACAO: "pap-secao-c",
    }
    icone = SECAO_ICONES.get(secao, "")
    titulo = html.escape(SECAO_LABELS[secao])
    cls = classes[secao]
    return (
        f'<div class="pap-secao-topo {cls}">'
        f'<div class="pap-secao-bar {cls}">'
        f'<span class="pap-secao-titulo">'
        f'<span class="pap-secao-ico">{icone}</span> {titulo}</span>'
        f"{_html_legenda_notas(compacta=True)}"
        f"</div></div>"
    )


def _html_barra_rodape() -> str:
    return (
        '<div class="pap-secao-topo pap-secao-rodape-wrap">'
        '<div class="pap-secao-bar pap-secao-rodape">'
        '<span class="pap-secao-titulo">Resumo final</span>'
        f"{_html_legenda_notas(compacta=True)}"
        "</div></div>"
    )


def _css_grelha_resumo() -> str:
    """Estilo da grelha de resumo alinhado ao mockup."""
    return f"""
<style>
section.main {{ background: #f1f5f9 !important; overflow-x: hidden !important; }}
section.main .block-container {{
    overflow-x: hidden !important; max-width: 100%; padding-top: 0.5rem;
}}

.pap-resumo-topo {{ margin-bottom: 0.5rem; }}
.pap-resumo-topo h1 {{ font-size: 1.85rem; font-weight: 800; color: #1e3a5f; margin: 0 0 0.35rem 0; }}
.pap-resumo-topo p {{ color: #64748b; margin: 0; font-size: 0.88rem; line-height: 1.4; }}

.pap-metricas {{
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px;
    margin: 1.1rem 0 1.35rem 0;
}}
.pap-metric-card {{
    display: flex; align-items: center; gap: 14px; background: #fff;
    border: 1px solid #e2e8f0; border-radius: 12px; padding: 16px 18px;
    box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06);
}}
.pap-metric-icone {{
    width: 42px; height: 42px; border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.1rem; flex-shrink: 0; font-weight: 700;
}}
.pap-icone-azul {{ background: #dbeafe; color: #2563eb; }}
.pap-icone-verde {{ background: #dcfce7; color: #16a34a; }}
.pap-icone-amarelo {{ background: #fef9c3; color: #ca8a04; }}
.pap-icone-roxo {{ background: #ede9fe; color: #7c3aed; }}
.pap-metric-label {{ color: #64748b; font-size: 0.78rem; font-weight: 500; }}
.pap-metric-num {{ font-weight: 800; color: #0f172a; font-size: 1.55rem; line-height: 1.1; margin: 2px 0; }}
.pap-metric-sub {{ color: #94a3b8; font-size: 0.74rem; }}

.pap-resumo-grid {{ margin: 0; }}
.pap-grelha-card {{
    background: transparent; border: none; box-shadow: none; overflow: visible;
    width: 100%;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) {{
    margin-bottom: 0 !important;
    padding-bottom: 0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) + div {{
    margin-top: 0 !important;
    padding-top: 0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) + div [data-testid="stAgGrid"] {{
    margin-top: 0 !important; padding-top: 0 !important;
    border-radius: 0 0 12px 12px !important;
    width: 100% !important;
    overflow: hidden !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) + div [data-testid="stAgGrid"] iframe {{
    overflow: hidden !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) + div {{
    width: 100% !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) + div .ag-root-wrapper {{
    border: 1px solid #bfdbfe !important; border-top: none !important;
    border-radius: 0 0 12px 12px !important;
    overflow: hidden !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) .pap-secao-topo {{
    border-radius: 12px 12px 0 0 !important;
    overflow: hidden !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo.pap-secao-a) .pap-secao-bar {{
    border: 1px solid #bfdbfe; border-bottom: 1px solid #e2e8f0;
    border-radius: 12px 12px 0 0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo.pap-secao-b) + div .ag-root-wrapper {{
    border-color: #bbf7d0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo.pap-secao-b) .pap-secao-bar {{
    border: 1px solid #bbf7d0; border-bottom: 1px solid #e2e8f0;
    border-radius: 12px 12px 0 0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo.pap-secao-c) + div .ag-root-wrapper {{
    border-color: #fde68a !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo.pap-secao-c) .pap-secao-bar {{
    border: 1px solid #fde68a; border-bottom: 1px solid #e2e8f0;
    border-radius: 12px 12px 0 0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-rodape-wrap) + div .ag-root-wrapper {{
    border-color: #e2e8f0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-rodape-wrap) .pap-secao-bar {{
    border: 1px solid #e2e8f0; border-bottom: 1px solid #e2e8f0;
    border-radius: 12px 12px 0 0 !important;
}}
section.main div[data-testid="stMarkdown"]:has(.pap-secao-topo) + div {{
    margin-bottom: 14px !important;
}}
.pap-secao-topo {{ margin: 0; padding: 0; border-radius: 12px 12px 0 0; overflow: hidden; }}

.pap-secao-bar {{
    display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 8px;
    border-bottom: 1px solid rgba(147, 197, 253, 0.35); padding: 11px 14px;
    text-align: left; font-weight: 700; font-size: 0.84rem; margin: 0;
    border-radius: 12px 12px 0 0;
}}
.pap-secao-titulo {{ flex: 1 1 auto; display: flex; align-items: center; gap: 8px; }}
.pap-secao-ico {{ font-size: 1rem; }}
.pap-secao-a {{ background: #eff6ff; color: #1d4ed8; }}
.pap-secao-b {{ background: #f0fdf4; color: #15803d; border-bottom-color: rgba(134, 239, 172, 0.45); }}
.pap-secao-c {{ background: #fffbeb; color: #b45309; border-bottom-color: rgba(252, 211, 77, 0.45); }}
.pap-secao-rodape {{ background: #f8fafc; color: #475569; border-bottom: 1px solid #e2e8f0; }}

.pap-hdr-pill[data-tip]:hover::after {{
    content: attr(data-tip); position: absolute; left: 50%; bottom: calc(100% + 8px);
    transform: translateX(-50%); background: #fff; color: #1e293b;
    border: 1px solid #e2e8f0; border-radius: 8px; padding: 7px 11px;
    font-size: 0.72rem; font-weight: 500; white-space: nowrap; z-index: 200;
    box-shadow: 0 4px 14px rgba(15, 23, 42, 0.12); pointer-events: none;
}}

.pap-legenda {{ display: flex; flex-wrap: wrap; gap: 10px; font-weight: 500; font-size: 0.72rem; color: #475569; }}
.pap-legenda-compacta {{ font-size: 0.68rem; gap: 8px; }}
.pap-legenda-centro {{ justify-content: center; margin: 14px 0 6px 0; }}
.pap-legenda-item {{ display: inline-flex; align-items: center; gap: 5px; white-space: nowrap; }}
.pap-legenda-cor {{
    display: inline-block; width: 10px; height: 10px; border-radius: 50%;
    border: 1px solid rgba(0,0,0,0.06);
}}
.pap-cor-muito-bom {{ background: {COR_CELULA_MUITO_BOM}; }}
.pap-cor-bom {{ background: {COR_CELULA_BOM}; }}
.pap-cor-suficiente {{ background: {COR_CELULA_SUFICIENTE}; }}
.pap-cor-insuficiente {{ background: {COR_CELULA_INSUFICIENTE}; }}
.pap-cor-vazia {{ background: #ffffff; border: 1px solid #cbd5e1; }}

.pap-resumo-export div[data-testid="stDownloadButton"] button {{
    background: #2563eb !important; color: #fff !important;
    border: 1px solid #1d4ed8 !important; border-radius: 10px !important;
    font-weight: 600 !important; padding: 0.55rem 1rem !important;
}}
.pap-resumo-export div[data-testid="stDownloadButton"] button:hover {{
    background: #1d4ed8 !important; border-color: #1e40af !important;
}}
</style>
<div class="pap-resumo-grid">
"""


def _guardar_nota_grelha(
    aluno_id: int,
    criterio: CriterioAvaliacao,
    nota: int,
    avaliacoes: dict,
) -> None:
    existente = avaliacoes.get(criterio)
    if criterio in CRITERIOS_MANUAIS:
        fonte = "manual"
    elif existente:
        fonte = (
            existente.fonte
            if "editado" in existente.fonte
            else f"{existente.fonte} (editado)"
        )
    else:
        fonte = "manual"
    storage.guardar_avaliacao(
        aluno_id,
        ResultadoCriterio(
            criterio=criterio,
            nota=nota,
            comentario=existente.comentario if existente else "",
            fonte=fonte,
        ),
    )
    st.session_state[f"nota_{criterio.value}_{aluno_id}"] = nota


def _percentagens_colunas(n_alunos: int) -> tuple[float, float]:
    largura_crit, largura_aluno = _larguras_colunas(n_alunos)
    total = largura_crit + n_alunos * largura_aluno
    return 100 * largura_crit / total, 100 * largura_aluno / total


def _config_colunas_grelha(
    colunas: list[tuple[str, AlunoRelatorio | None]],
) -> tuple[list[str], list[str]]:
    colunas_ordem = ["Critério"]
    colunas_desativadas: list[str] = ["Critério"]

    for indice, (_nome, aluno) in enumerate(colunas):
        chave = _chave_coluna_grelha(aluno, indice)
        colunas_ordem.append(chave)
        if aluno is None:
            colunas_desativadas.append(chave)
    return colunas_ordem, colunas_desativadas


def _avaliacoes_colunas(
    colunas: list[tuple[str, AlunoRelatorio | None]],
) -> dict[int, dict]:
    ids = [a.id for _, a in colunas if a is not None]
    return {aid: storage.obter_avaliacoes(aid) for aid in ids}


def _linhas_secao(
    colunas: list[tuple[str, AlunoRelatorio | None]],
    secao: SecaoAvaliacao,
    avaliacoes: dict[int, dict] | None = None,
) -> tuple[list[dict], dict[str, CriterioAvaliacao]]:
    if avaliacoes is None:
        avaliacoes = _avaliacoes_colunas(colunas)
    chaves = [_chave_coluna_grelha(a, i) for i, (_, a) in enumerate(colunas)]
    mapa_itens: dict[str, CriterioAvaliacao] = {}
    linhas: list[dict] = []

    for criterio in CRITERIOS_POR_SECAO[secao]:
        rotulo = _rotulo_criterio_resumo(criterio)
        mapa_itens[rotulo] = criterio
        linha = {"Critério": rotulo}
        for (_, aluno), chave in zip(colunas, chaves):
            if aluno is None:
                linha[chave] = None
                continue
            av = avaliacoes[aluno.id]
            linha[chave] = int(av[criterio].nota) if criterio in av else None
        linhas.append(linha)

    linha_tot = {"Critério": "Total:"}
    for (_, aluno), chave in zip(colunas, chaves):
        if aluno is None:
            linha_tot[chave] = None
            continue
        media = media_secao_arredondada(avaliacoes[aluno.id], secao)
        linha_tot[chave] = media
    linhas.append(linha_tot)

    return linhas, mapa_itens


def _linhas_rodape(
    colunas: list[tuple[str, AlunoRelatorio | None]],
    avaliacoes: dict[int, dict] | None = None,
) -> tuple[list[dict], dict[str, CriterioAvaliacao]]:
    if avaliacoes is None:
        avaliacoes = _avaliacoes_colunas(colunas)
    chaves = [_chave_coluna_grelha(a, i) for i, (_, a) in enumerate(colunas)]
    mapa_itens: dict[str, CriterioAvaliacao] = {}
    linhas: list[dict] = []

    linha_pond: dict = {"Critério": "Média ponderada"}
    for (_, aluno), chave in zip(colunas, chaves):
        if aluno is None:
            linha_pond[chave] = None
            continue
        final = avaliacao_final(avaliacoes[aluno.id])
        linha_pond[chave] = final
    linhas.append(linha_pond)

    linha_nf: dict = {"Critério": "Nota Final"}
    for (_, aluno), chave in zip(colunas, chaves):
        if aluno is None:
            linha_nf[chave] = None
            continue
        linha_nf[chave] = nota_final_arredondada(avaliacoes[aluno.id])
    linhas.append(linha_nf)

    rotulo_auto = _rotulo_criterio_resumo(CriterioAvaliacao.AUTOAVALIACAO)
    mapa_itens[rotulo_auto] = CriterioAvaliacao.AUTOAVALIACAO
    linha_auto: dict = {"Critério": rotulo_auto}
    for (_, aluno), chave in zip(colunas, chaves):
        if aluno is None:
            linha_auto[chave] = None
            continue
        av = avaliacoes[aluno.id]
        c = CriterioAvaliacao.AUTOAVALIACAO
        linha_auto[chave] = int(av[c].nota) if c in av else None
    linhas.append(linha_auto)

    return linhas, mapa_itens


def _aplicar_edicao_aggrid(
    df: pd.DataFrame,
    colunas: list[tuple[str, AlunoRelatorio | None]],
    mapa_itens: dict[str, CriterioAvaliacao],
) -> bool:
    chaves = [_chave_coluna_grelha(a, i) for i, (_, a) in enumerate(colunas)]
    mudou = False
    for _, linha in df.iterrows():
        rotulo = linha.get("Critério", "")
        if rotulo in _ROTULOS_LINHA_SO_LEITURA:
            continue
        criterio = mapa_itens.get(rotulo)
        if not criterio:
            continue
        for (_, aluno), chave in zip(colunas, chaves):
            if aluno is None:
                continue
            novo = linha.get(chave)
            if novo is None or (isinstance(novo, float) and pd.isna(novo)):
                continue
            try:
                nota_nova = int(round(float(novo)))
            except (TypeError, ValueError):
                continue
            if not 1 <= nota_nova <= NOTA_MAXIMA:
                continue
            av = storage.obter_avaliacoes(aluno.id)
            atual = int(av[criterio].nota) if criterio in av else None
            if nota_nova == atual:
                continue
            _guardar_nota_grelha(aluno.id, criterio, nota_nova, av)
            mudou = True
    return mudou


def _renderizar_editor_grelha(
    linhas: list[dict],
    colunas: list[tuple[str, AlunoRelatorio | None]],
    colunas_ordem: list[str],
    editor_key: str,
    mapa_itens: dict[str, CriterioAvaliacao],
) -> None:
    df = pd.DataFrame(linhas)[colunas_ordem]
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        editable=_JS_CELULA_EDITAVEL,
        sortable=False,
        filterable=False,
        resizable=False,
        suppressMenu=True,
        suppressSizeToFit=True,
        flex=1,
        minWidth=LARGURA_MIN_ALUNO,
        cellStyle=_JS_ESTILO_CELULA,
        valueFormatter=_JS_FORMATO_CELULA,
    )
    gb.configure_column(
        "Critério",
        editable=False,
        pinned="left",
        width=LARGURA_CRITERIO,
        minWidth=LARGURA_CRITERIO,
        maxWidth=LARGURA_CRITERIO,
        flex=0,
        suppressSizeToFit=True,
        cellStyle=_JS_ESTILO_CELULA,
        headerComponent=_JS_HEADER_CRITERIO,
    )
    for indice, (nome, aluno) in enumerate(colunas):
        chave = _chave_coluna_grelha(aluno, indice)
        if chave not in colunas_ordem:
            continue
        gb.configure_column(
            chave,
            headerName=_iniciais_nome(nome),
            headerTooltip=nome,
            headerComponent=_JS_HEADER_ALUNO,
            editable=_JS_CELULA_EDITAVEL if aluno is not None else False,
            flex=1,
            minWidth=LARGURA_MIN_ALUNO,
            suppressSizeToFit=True,
            cellStyle=_JS_ESTILO_CELULA,
            valueFormatter=_JS_FORMATO_CELULA,
        )
    gb.configure_grid_options(
        domLayout="normal",
        suppressHorizontalScroll=True,
        alwaysShowHorizontalScroll=False,
        alwaysShowVerticalScroll=False,
        stopEditingWhenCellsLoseFocus=True,
        singleClickEdit=True,
        rowHeight=ALTURA_LINHA_GRELHA,
        headerHeight=ALTURA_CABECALHO_GRELHA,
        suppressColumnVirtualisation=True,
        suppressAutoSize=True,
        getRowClass=_JS_LINHA_TOTAL,
    )

    resposta = AgGrid(
        df,
        gridOptions=_fixar_opcoes_grelha(gb),
        data_return_mode=DataReturnMode.AS_INPUT,
        update_on=["cellValueChanged"],
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        height=_altura_grelha(len(linhas)),
        theme="balham",
        key=editor_key,
        server_sync_strategy="server_wins",
        custom_css=_CSS_AGGRID_RESUMO,
    )
    if resposta["data"] is None:
        return
    if _aplicar_edicao_aggrid(
        pd.DataFrame(resposta["data"]), colunas, mapa_itens
    ):
        st.rerun()


def _limpar_session_notas_criterios(
    aluno_ids: list[int],
    criterios: list[CriterioAvaliacao],
) -> None:
    for aluno_id in aluno_ids:
        for criterio in criterios:
            st.session_state.pop(f"nota_{criterio.value}_{aluno_id}", None)
            st.session_state.pop(f"com_{criterio.value}_{aluno_id}", None)


def _ids_alunos_colunas(
    colunas: list[tuple[str, AlunoRelatorio | None]],
) -> list[int]:
    return [aluno.id for _, aluno in colunas if aluno is not None]


def _limpar_dados_secao_resumo(
    secao: SecaoAvaliacao,
    colunas: list[tuple[str, AlunoRelatorio | None]],
) -> int:
    aluno_ids = _ids_alunos_colunas(colunas)
    criterios = CRITERIOS_POR_SECAO[secao]
    removidos = storage.apagar_avaliacoes_criterios(aluno_ids, criterios)
    _limpar_session_notas_criterios(aluno_ids, criterios)
    st.session_state.pop("_acta_bytes", None)
    return removidos


def _botao_limpar_tabela_resumo(chave: str, limpar_fn) -> None:
    _, col_btn = st.columns([5, 1])
    with col_btn:
        if st.button(
            "Limpar notas",
            key=f"limpar_{chave}",
            use_container_width=True,
        ):
            removidos = limpar_fn()
            if removidos:
                st.toast(f"{removidos} nota(s) removida(s).")
            else:
                st.toast("Não havia notas para limpar.")
            st.rerun()


def _pagina_resumo(alunos: list[AlunoRelatorio]) -> None:
    if not alunos:
        st.info("Importe relatórios na barra lateral.")
        return

    colunas = colunas_turma_ordenadas(alunos)
    avaliacoes = _avaliacoes_colunas(colunas)
    metricas = _metricas_resumo_turma(colunas, avaliacoes)
    disp, _motor = _ia_status()

    topo_esq, topo_dir = st.columns([5, 1])
    with topo_esq:
        st.markdown(
            '<p style="color: #64748b; margin: 0 0 0.75rem 0; font-size: 0.88rem; line-height: 1.4;">'
            "Edita as notas diretamente na grelha — guarda ao sair da célula. "
            "Totais e nota final recalculam automaticamente.</p>",
            unsafe_allow_html=True,
        )
    with topo_dir:
        if not disp:
            st.warning("IA indisponível")
        elif st.button(
            "📋 Avaliar todos os relatórios",
            type="primary",
            use_container_width=True,
        ):
            _avaliar_todos(alunos)
            st.rerun()

    st.markdown(_html_metricas_resumo(metricas), unsafe_allow_html=True)

    colunas_ordem, _colunas_desativadas = _config_colunas_grelha(colunas)

    st.markdown(_css_grelha_resumo(), unsafe_allow_html=True)
    st.markdown('<div class="pap-grelha-card">', unsafe_allow_html=True)

    for secao in SecaoAvaliacao:
        st.markdown(_html_barra_secao(secao), unsafe_allow_html=True)
        linhas, mapa = _linhas_secao(colunas, secao, avaliacoes)
        _renderizar_editor_grelha(
            linhas,
            colunas,
            colunas_ordem,
            f"resumo_{secao.value}",
            mapa,
        )
        _botao_limpar_tabela_resumo(
            secao.value,
            lambda secao=secao: _limpar_dados_secao_resumo(secao, colunas),
        )

    st.markdown(_html_barra_rodape(), unsafe_allow_html=True)
    linhas_rodape, mapa_rodape = _linhas_rodape(colunas, avaliacoes)
    _renderizar_editor_grelha(
        linhas_rodape,
        colunas,
        colunas_ordem,
        "resumo_rodape",
        mapa_rodape,
    )

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    _, col_export = st.columns([5, 1])
    with col_export:
        st.markdown('<div class="pap-resumo-export">', unsafe_allow_html=True)
        if st.button("⬇ Exportar Excel", use_container_width=True, key="btn_export_acta"):
            try:
                alunos_export = [a for _, a in colunas if a is not None]
                resultado = sincronizar_acta(alunos_export, avaliacoes)
                st.session_state["_acta_bytes"] = resultado.bytes_ficheiro
                partes = []
                if resultado.exportados:
                    partes.append(f"{resultado.exportados} notas exportadas para a Acta")
                for aviso in resultado.avisos:
                    partes.append(aviso)
                if partes:
                    st.session_state["_acta_msg"] = " · ".join(partes)
                else:
                    st.session_state["_acta_msg"] = "Nada a exportar — Acta inalterada."
                st.rerun()
            except Exception as exc:
                st.error(f"Erro ao sincronizar Acta: {exc}")

        if msg := st.session_state.pop("_acta_msg", None):
            st.success(msg)

        acta_bytes = st.session_state.get("_acta_bytes")
        if acta_bytes:
            st.download_button(
                "Descarregar Acta",
                data=acta_bytes,
                file_name="Acta_Pap2526.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_acta",
            )
        st.markdown("</div>", unsafe_allow_html=True)


_inicializar_instrucoes()

_tentar_restaurar_sessao_auth()

sessao = _sessao_auth()
if not sessao:
    _pagina_login()
    st.stop()

_renderizar_titulo_plataforma()

pagina = st.radio(
    "Nav",
    ["Resumo", "Relatórios", "Definições"],
    horizontal=True,
    label_visibility="collapsed",
)
_renderizar_subtitulo_pagina(pagina)

with st.sidebar:
    st.caption(f"**{sessao['nome']}**")
    st.caption("Administrador" if sessao["role"] == "admin" else "Professor")
    if st.button("Terminar sessão", use_container_width=True):
        _terminar_sessao_auth()
        st.rerun()
    st.divider()
    st.header("Importar Relatórios")
    uploads = st.file_uploader("Extensão (.docx)", type=["docx"], accept_multiple_files=True)
    if uploads and st.button("Importar", type="primary"):
        st.success(f"{_importar_ficheiros(uploads)} importado(s).")
        st.rerun()
    st.divider()
    alunos = storage.listar_alunos()
    st.metric("Alunos", len(alunos))
    disp, motor = _ia_status()
    if disp:
        st.success(f"IA: {motor}")
    else:
        st.warning(f"IA indisponível — {motor}")
    if st.button("Verificar motor IA", use_container_width=True, key="btn_verificar_ia"):
        st.session_state.pop("_ia_status", None)
        st.rerun()
    st.caption(descricao_armazenamento())
    st.caption("Use iniciar.bat na Drive da escola.")

alunos = storage.listar_alunos()
if pagina == "Definições":
    _pagina_configuracao()
elif pagina == "Resumo":
    _pagina_resumo(alunos)
elif not alunos:
    st.info("Importe relatórios .docx na barra lateral.")
else:
    st.caption(f"{len(alunos)} alunos — expande cada um para avaliar ou ver detalhes.")
    for aluno in alunos:
        av = storage.obter_avaliacoes(aluno.id)
        media_rel = media_secao_arredondada(av, SecaoAvaliacao.RELATORIO)
        media_txt = f" — {media_rel} val." if media_rel is not None else ""
        rotulo = f"{aluno.nome}{media_txt}"
        with st.expander(rotulo, expanded=False):
            _renderizar_tab_aluno(aluno)
